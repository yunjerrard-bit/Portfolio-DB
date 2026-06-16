"""Phase 4 Plan 02 (04-02) Task 1: OUT-04 frozen panes openpyxl 회귀 테스트.

검증 대상 (이미 구현됨 — 04-02 must_haves):
  - 모든 시트의 행 1~5 가 frozen (행 6 이 첫 비고정 셀 → freeze_panes 가 "6" 으로 끝남)
  - 시트1: 행 1~5 + A열 고정 (freeze_panes == "B6")
  - 종목 시트: 행 1~5 + A열(날짜) 고정 (freeze_panes == "B6")

XlsxWriter `ws.freeze_panes(row, col)` 는 "첫 비고정 셀" 규약:
  - sheet_per_ticker.py: freeze_panes(5, 1) → openpyxl 읽기 시 "B6"
  - sheet_portfolio.py:  freeze_panes(5, 1) → openpyxl 읽기 시 "B6"

후속 phase 가 frozen panes 를 깨뜨리면 이 회귀 테스트가 즉시 적색이 된다.

fixture (env_file / mock_pipeline_env) 는 tests/test_smoke_n_tickers.py 패턴을
차용 — fetch_ohlcv stub + 캐시 디렉토리 tmp_path 격리 (네트워크 없음, 결정론적).
"""

from __future__ import annotations

from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
import pytest

import stocksig.io.cache as cache_mod
import stocksig.io.market as market_mod
from stocksig.main_run import run


# --- helpers ---------------------------------------------------------------


def _make_ohlcv(rows: int = 2700, seed: int = 42) -> pd.DataFrame:
    """결정론적 OHLCV — test_smoke_n_tickers._make_ohlcv 와 동일 구조."""
    rng = np.random.default_rng(seed=seed)
    dates = pd.date_range(end=pd.Timestamp("2026-05-20"), periods=rows, freq="B")
    drift = rng.normal(loc=0.0, scale=1.0, size=rows).cumsum() * 0.1
    close = 100.0 + drift
    df = pd.DataFrame(
        {
            "Open": close + rng.normal(0.0, 0.5, rows),
            "High": close * 1.02,
            "Low": close * 0.98,
            "Close": close,
            "Volume": rng.integers(1_000_000, 10_000_000, rows),
        },
        index=dates,
    )
    df.index.name = "Date"
    return df


@pytest.fixture
def env_file(tmp_path: Path) -> Path:
    p = tmp_path / ".env"
    p.write_text(
        "EDGAR_USER_AGENT_EMAIL=test@example.com\nOPENDART_API_KEY=test-key\n",
        encoding="utf-8",
    )
    return p


@pytest.fixture
def mock_pipeline_env(monkeypatch, tmp_path):
    """fetch_ohlcv stub + 기업명 fetch stub + cache dir 격리 (test_smoke_n_tickers 패턴 차용).

    06-01 (Pitfall 5): run() 경로에 추가된 fetch_company_name 이 실제 yfinance .info 를
    치지 않도록 결정론적 stub(티커 반환)으로 격리하고, company 캐시 디렉토리도 tmp 격리한다.
    freeze 단언 "B6" 는 기업명(B 비고정)이 추가돼도 불변이어야 한다.
    """
    cache_dir = tmp_path / ".cache" / "ohlcv"
    monkeypatch.setattr(cache_mod, "_DEFAULT_DIR", cache_dir)
    monkeypatch.setattr(cache_mod, "_cache", None)
    # company 캐시 격리 (운영 .cache/company 오염 방지)
    monkeypatch.setattr(cache_mod, "_NAME_DIR", tmp_path / ".cache" / "company")
    monkeypatch.setattr(cache_mod, "_name_cache", None)

    def _stub_fetch(ticker: str) -> pd.DataFrame:
        seed = abs(hash(ticker)) % 10_000
        return _make_ohlcv(rows=2700, seed=seed)

    monkeypatch.setattr(market_mod, "fetch_ohlcv", _stub_fetch)
    # 기업명 fetch stub — 티커=이름 (네트워크 없음). main_run 의 바인딩을 patch.
    import stocksig.main_run as main_run_mod

    monkeypatch.setattr(main_run_mod, "fetch_company_name", lambda t: t)
    yield
    # Windows: 캐시 파일 핸들 close 로 tmp_path 정리 실패 방지 (test_cache.py:25-33 패턴)
    for attr in ("_cache", "_name_cache"):
        c = getattr(cache_mod, attr, None)
        if c is not None:
            try:
                c.close()
            except Exception:
                pass
            monkeypatch.setattr(cache_mod, attr, None)


# --- tests -----------------------------------------------------------------


def test_all_sheets_freeze_rows_1_to_5(mock_pipeline_env, tmp_path, env_file):
    """OUT-04: 모든 시트의 행 1~5 가 frozen (freeze_panes 가 행 6 = "6" 으로 끝남)."""
    tickers = tmp_path / "tickers.txt"
    tickers.write_text("AAPL\nMSFT\n", encoding="utf-8")

    out = run(tickers, env_file, tmp_path / "output")

    wb = openpyxl.load_workbook(out)
    for name in wb.sheetnames:
        fp = wb[name].freeze_panes
        assert fp is not None, f"시트 {name!r} freeze_panes 가 None (고정 없음)"
        assert fp.endswith("6"), (
            f"시트 {name!r} freeze_panes={fp!r} — 행 6 이 첫 비고정 셀이어야 "
            "(행 1~5 고정)"
        )


def test_portfolio_sheet_freezes_rows_and_col_a(
    mock_pipeline_env, tmp_path, env_file
):
    """시트1: 행 1~5 + A열 고정 → freeze_panes == "B6"."""
    tickers = tmp_path / "tickers.txt"
    tickers.write_text("AAPL\nMSFT\n", encoding="utf-8")

    out = run(tickers, env_file, tmp_path / "output")

    wb = openpyxl.load_workbook(out)
    assert wb["시트1"].freeze_panes == "B6", (
        f"시트1 freeze_panes={wb['시트1'].freeze_panes!r} (기대 'B6')"
    )


def test_per_ticker_sheet_freezes_rows_and_col_a(
    mock_pipeline_env, tmp_path, env_file
):
    """종목 시트: 행 1~5 + A열(날짜) 고정 → freeze_panes == "B6"."""
    tickers = tmp_path / "tickers.txt"
    tickers.write_text("AAPL\nMSFT\n", encoding="utf-8")

    out = run(tickers, env_file, tmp_path / "output")

    wb = openpyxl.load_workbook(out)
    assert wb["AAPL"].freeze_panes == "B6", (
        f"AAPL 시트 freeze_panes={wb['AAPL'].freeze_panes!r} (기대 'B6')"
    )
