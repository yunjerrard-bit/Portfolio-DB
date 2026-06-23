"""FUND-07/08 통합 테스트 — main_run.run() 히스토리 경로 배선 (Plan 07-04, D-07).

검증 대상:
  - test_run_creates_history_db: run() 1회(네트워크 0 mock) → 격리 data/fundamentals.db
    raw_facts 행 누적(count_rows > 0). 델타 감지 → full-fetch 경로(SC1 누적).
  - test_steady_state_history_zero_full_fetch (SC3): delta_state 전 종목 시드 + probe 가
    동일 accession 반환 → run() 후 fetch_*_quarterly_raw 호출 0·full_fetch == 0.
  - test_sheet1_unchanged_by_history (D-07): 히스토리 경로 mock on/off 두 실행의
    write_portfolio_sheet 인자(results/failures/input_order)와 산출 xlsx 시트1 셀을
    스냅샷 비교해 바이트/셀 단위 불변임을 직접 단언 — 막연한 회귀 통과에 의존하지 않음.
  - test_cache_dir_unchanged_by_history (SC5): 히스토리 실행 전후 `.cache/` 파일 목록
    (이름·개수) 불변 — 히스토리는 data/fundamentals.db 만 쓰고 .cache/fundamentals
    7일 캐시는 건드리지 않는다(data/ 와 .cache/ 분리).
  - test_history_failure_does_not_break_sheet1: sync_ticker_history 예외 → run() 이
    portfolio_*.xlsx 정상 저장(시트1 보호, D-07 분리 보장, T-07-11).

모든 외부 호출은 mock(네트워크 0): OHLCV/시트1 펀더멘털·기업명은 smoke 패턴 stub,
히스토리 경로는 probe_edgar_accession / fetch_edgar_quarterly_raw 를 mock.
운영 DB·캐시는 conftest `_isolated_fundamentals_db`/`_isolated_disk_cache` 로 tmp 격리.
"""

from __future__ import annotations

import os
from pathlib import Path

import openpyxl
import pandas as pd

from stocksig.io import dart_client
from stocksig.io import edgar_client
from stocksig.io import fundamentals
from stocksig.io import fundamentals_delta as fd
from stocksig.io import fundamentals_store as fs
from stocksig.main import run

# 추출기 11-key dict 행 (Plan 02 계약 — delta 가 12-tuple 로 변환해 upsert).
_FETCH_ROW = {
    "ticker": "AAPL",
    "source": "EDGAR",
    "quarter": "2026Q1",
    "field": "revenue",
    "value": 1000.0,
    "unit": "USD",
    "accession": "ACC2",
    "period_start": "2026-01-01",
    "period_end": "2026-03-31",
    "period_type": "duration",
    "reprt_code": None,
}


def _setup_mock_yfinance(mocker, df: pd.DataFrame) -> None:
    """`yf.Ticker(...).history(...)` + 기업명 fetch 를 결정론적 stub (네트워크 0).

    test_smoke_end_to_end._setup_mock_yfinance 와 동일 패턴.
    """
    ticker_class = mocker.patch("stocksig.io.market.yf.Ticker")
    ticker_class.return_value.history.return_value = df
    mocker.patch("stocksig.io.company.fetch_company_name", side_effect=lambda t: t)
    mocker.patch("stocksig.main_run.fetch_company_name", side_effect=lambda t: t)


def _disable_history(mocker) -> None:
    """히스토리 경로 외부 호출을 전부 무력화 (probe None → 즉시 SKIP, fetch 미호출)."""
    mocker.patch.object(fd, "probe_edgar_accession", return_value=None)
    mocker.patch.object(fd, "probe_dart_rcept", return_value=None)


def test_run_creates_history_db(
    mocker, mock_ohlcv_df, tmp_path, tmp_tickers_file, tmp_env_file
):
    """SC1: run() 1회 → 델타 감지(state 부재) → full-fetch → data/fundamentals.db 누적."""
    _setup_mock_yfinance(mocker, mock_ohlcv_df)
    # 히스토리 probe 가 accession 반환(델타 있음, state 부재) → full-fetch 경로.
    mocker.patch.object(fd, "probe_edgar_accession", return_value="ACC1")
    row = dict(_FETCH_ROW, accession="ACC1")
    mocker.patch.object(
        edgar_client, "fetch_edgar_quarterly_raw", return_value=[row]
    )
    tickers = tmp_tickers_file("AAPL\n")
    env = tmp_env_file(
        "EDGAR_USER_AGENT_EMAIL=test@example.com\nOPENDART_API_KEY=test-key\n"
    )

    output_path = run(tickers, env, tmp_path / "output")

    assert output_path.exists()  # 시트1 산출물 정상
    assert fs.count_rows("AAPL") > 0  # 히스토리 raw 누적 (SC1)
    assert fs.get_last_accession("AAPL", "EDGAR") == "ACC1"  # forward 누적
    stats = fs.get_delta_stats()
    assert stats["full_fetch"] == 1


def test_steady_state_history_zero_full_fetch(
    mocker, mock_ohlcv_df, tmp_path, tmp_tickers_file, tmp_env_file
):
    """SC3: delta_state 시드 + probe 동일 accession → run() 후 full-fetch 호출 0."""
    _setup_mock_yfinance(mocker, mock_ohlcv_df)
    # 전 종목 delta_state 시드 후 probe 가 같은 값 반환 → 전부 SKIP.
    fs.set_last_accession("AAPL", "EDGAR", "ACC-SAME")
    fs.set_last_accession("MSFT", "EDGAR", "ACC-SAME")
    mocker.patch.object(fd, "probe_edgar_accession", return_value="ACC-SAME")
    spy = mocker.spy(edgar_client, "fetch_edgar_quarterly_raw")
    tickers = tmp_tickers_file("AAPL\nMSFT\n")
    env = tmp_env_file(
        "EDGAR_USER_AGENT_EMAIL=test@example.com\nOPENDART_API_KEY=test-key\n"
    )

    output_path = run(tickers, env, tmp_path / "output")

    assert output_path.exists()
    assert spy.call_count == 0  # 평소 실행 외부 전체호출 0 (SC3)
    stats = fs.get_delta_stats()
    assert stats["full_fetch"] == 0
    assert stats["delta_hit"] == 2  # 두 종목 모두 델타 없음


def _portfolio_call_snapshot(spy) -> tuple:
    """write_portfolio_sheet spy 호출에서 시트1 신호 인자(results/failures/input_order) 추출.

    results 의 셀/색 신호를 결정하는 (symbol, scalars, fundamentals) 핵심을 비교 가능한
    형태로 직렬화 — wb/formats(객체 동일성 무의미)는 제외.
    """
    assert spy.call_count == 1
    _wb, _formats, results, failures, input_order = spy.call_args.args[:5]
    res_keys = tuple(
        (r.spec.symbol, tuple(sorted(r.enriched_df.attrs.get("scalars", {}).items())))
        for r in results
    )
    fail_keys = tuple(f.spec.symbol for f in failures)
    return (res_keys, fail_keys, tuple(input_order))


def test_sheet1_unchanged_by_history(
    mocker, mock_ohlcv_df, tmp_path, tmp_tickers_file, tmp_env_file
):
    """D-07: 히스토리 mock on/off 두 실행의 시트1(write_portfolio_sheet 인자·셀) 불변 직접 단언."""
    tickers = tmp_tickers_file("AAPL\n")
    env = tmp_env_file(
        "EDGAR_USER_AGENT_EMAIL=test@example.com\nOPENDART_API_KEY=test-key\n"
    )

    # --- 실행 A: 히스토리 경로 OFF (probe None → 아무 누적도 안 함) ---
    _setup_mock_yfinance(mocker, mock_ohlcv_df)
    _disable_history(mocker)
    spy_off = mocker.spy(
        __import__("stocksig.main_run", fromlist=["write_portfolio_sheet"]),
        "write_portfolio_sheet",
    )
    out_off = run(tickers, env, tmp_path / "out_off")
    snap_off = _portfolio_call_snapshot(spy_off)
    wb_off = openpyxl.load_workbook(out_off)
    ws_off = wb_off["시트1"]
    cells_off = [
        (c.value, c.font.color.rgb if c.font.color else None)
        for row in ws_off.iter_rows(min_row=5, max_row=ws_off.max_row)
        for c in row
    ]

    mocker.stopall()

    # --- 실행 B: 히스토리 경로 ON (델타 감지·full-fetch 실제 발생) ---
    _setup_mock_yfinance(mocker, mock_ohlcv_df)
    mocker.patch.object(fd, "probe_edgar_accession", return_value="ACC1")
    mocker.patch.object(
        edgar_client,
        "fetch_edgar_quarterly_raw",
        return_value=[dict(_FETCH_ROW, accession="ACC1")],
    )
    spy_on = mocker.spy(
        __import__("stocksig.main_run", fromlist=["write_portfolio_sheet"]),
        "write_portfolio_sheet",
    )
    out_on = run(tickers, env, tmp_path / "out_on")
    snap_on = _portfolio_call_snapshot(spy_on)
    wb_on = openpyxl.load_workbook(out_on)
    ws_on = wb_on["시트1"]
    cells_on = [
        (c.value, c.font.color.rgb if c.font.color else None)
        for row in ws_on.iter_rows(min_row=5, max_row=ws_on.max_row)
        for c in row
    ]

    # 히스토리가 실제로 누적되었음(ON 실행이 OFF 와 진짜 다른 경로였음)을 확인.
    assert fs.count_rows("AAPL") > 0
    # 그럼에도 시트1 write 인자·셀(값·색 신호)은 바이트/셀 단위로 동일 (D-07 회귀 0).
    assert snap_on == snap_off
    assert cells_on == cells_off


def test_cache_dir_unchanged_by_history(
    mocker, mock_ohlcv_df, tmp_path, tmp_tickers_file, tmp_env_file
):
    """SC5: 히스토리 실행 전후 `.cache/` 파일 목록 불변 (data/ 와 .cache/ 분리)."""
    _setup_mock_yfinance(mocker, mock_ohlcv_df)
    mocker.patch.object(fd, "probe_edgar_accession", return_value="ACC1")
    mocker.patch.object(
        edgar_client,
        "fetch_edgar_quarterly_raw",
        return_value=[dict(_FETCH_ROW, accession="ACC1")],
    )
    tickers = tmp_tickers_file("AAPL\n")
    env = tmp_env_file(
        "EDGAR_USER_AGENT_EMAIL=test@example.com\nOPENDART_API_KEY=test-key\n"
    )

    # conftest _isolated_disk_cache 가 .cache 를 tmp_path/.cache 로 격리.
    cache_root = tmp_path / ".cache"

    def _listing() -> set[str]:
        if not cache_root.exists():
            return set()
        return {
            os.path.relpath(os.path.join(dp, f), cache_root)
            for dp, _dn, fns in os.walk(cache_root)
            for f in fns
        }

    # 1차 run() 으로 시트1 OHLCV/펀더멘털 캐시를 채운다 (베이스라인).
    run(tickers, env, tmp_path / "out1")
    before = _listing()

    # 히스토리 카운터만 리셋되도록 다시 run() — 같은 날 캐시는 HIT.
    run(tickers, env, tmp_path / "out2")
    after = _listing()

    assert fs.count_rows("AAPL") > 0  # 히스토리는 data/fundamentals.db 에 누적
    # 히스토리 경로는 .cache/ 를 건드리지 않는다 — 파일 목록 불변.
    assert after == before


def test_history_failure_does_not_break_sheet1(
    mocker, mock_ohlcv_df, tmp_path, tmp_tickers_file, tmp_env_file
):
    """T-07-11: sync_ticker_history 예외 → run() 이 portfolio_*.xlsx 정상 저장 (시트1 보호)."""
    _setup_mock_yfinance(mocker, mock_ohlcv_df)
    mocker.patch.object(
        fd,
        "sync_ticker_history",
        side_effect=RuntimeError("히스토리 경로 폭발"),
    )
    tickers = tmp_tickers_file("AAPL\n")
    env = tmp_env_file(
        "EDGAR_USER_AGENT_EMAIL=test@example.com\nOPENDART_API_KEY=test-key\n"
    )

    output_path = run(tickers, env, tmp_path / "output")

    # 히스토리 경로가 던져도 시트1 산출물은 정상 (D-07 분리, 예외 흡수).
    assert output_path.exists()
    wb = openpyxl.load_workbook(output_path)
    assert "시트1" in wb.sheetnames
    assert "AAPL" in wb.sheetnames


# --- Plan 10-02 Task 2: 단일 원천·외부 호출 0 (FUND-11) ----------------------

# 12-tuple raw_facts 행 = (ticker, source, quarter, field, value, unit,
#   accession, period_start, period_end, period_type, reprt_code, fetched_at).
# upsert_quarters 가 받는 형태(fundamentals_store.upsert_quarters docstring L126).
def _store_row(quarter: str, field: str, value: float, period_type: str) -> tuple:
    return (
        "AAPL", "EDGAR", quarter, field, value, "USD",
        "ACC-SEED", "2026-01-01", "2026-03-31", period_type, None, "2026-01-01T00:00:00",
    )


def _seed_store_aapl() -> None:
    """compute_matrix 가 PER/GPM/OPM 최신열을 산출하도록 격리 store 에 raw 5분기 적재.

    history_fixtures._series_rows 와 동일 지표 구성(revenue/gross_profit/operating_income/
    net_income/total_equity/total_assets/shares_outstanding)을 12-tuple 로 직접 upsert —
    compute_matrix("AAPL") 가 외부 호출 0(SQLite SELECT)로 매트릭스를 만든다.
    """
    quarters = ["2025Q1", "2025Q2", "2025Q3", "2025Q4", "2026Q1"]
    rows: list[tuple] = []
    for i, q in enumerate(quarters):
        scale = 1.0 + 0.1 * i
        rows += [
            _store_row(q, "revenue", 1000.0 * scale, "duration"),
            _store_row(q, "gross_profit", 400.0 * scale, "duration"),
            # OPM 은 REGISTRY 상 store field 명이 "op_income" (metrics_registry.py:82).
            _store_row(q, "op_income", 300.0 * scale, "duration"),
            _store_row(q, "net_income", 250.0 * scale, "duration"),
            _store_row(q, "total_equity", 2000.0 * scale, "instant"),
            _store_row(q, "total_assets", 5000.0 * scale, "instant"),
            _store_row(q, "shares_outstanding", 100.0, "instant"),
        ]
    fs.upsert_quarters(rows)


def test_run_no_legacy_fetch(
    mocker, mock_ohlcv_df, tmp_path, tmp_tickers_file, tmp_env_file
):
    """FUND-11(위생): run() 시트1 경로에서 구 fetch(fetch_fundamentals/edgar_cached/
    dart_cached) 호출 0 — 단일 원천(store/registry)으로 이관됨."""
    _setup_mock_yfinance(mocker, mock_ohlcv_df)
    _disable_history(mocker)  # SYNC 외부 호출 차단 (probe None → SKIP)

    spy_fund = mocker.spy(fundamentals, "fetch_fundamentals")
    spy_edgar = mocker.spy(edgar_client, "fetch_edgar_cached")
    spy_dart = mocker.spy(dart_client, "fetch_dart_cached")

    tickers = tmp_tickers_file("AAPL\n")
    env = tmp_env_file(
        "EDGAR_USER_AGENT_EMAIL=test@example.com\nOPENDART_API_KEY=test-key\n"
    )

    output_path = run(tickers, env, tmp_path / "output")

    assert output_path.exists()
    # 시트1 경로 구 fetch 미호출 (단일 원천 — store/registry 읽기로 대체).
    assert spy_fund.call_count == 0
    assert spy_edgar.call_count == 0
    assert spy_dart.call_count == 0


def test_run_single_source(
    mocker, mock_ohlcv_df, tmp_path, tmp_tickers_file, tmp_env_file
):
    """FUND-11: sync(DB 적재) 후 시트1 펀더멘털이 store 매트릭스 최신열에서 채워진다.

    격리 store 에 raw 사전 적재 → run() 의 READ 단계가 compute_matrix(SQLite SELECT)로
    읽어 res.fundamentals 를 store 값으로 재할당. 외부 펀더멘털 fetch 호출 0.
    """
    _setup_mock_yfinance(mocker, mock_ohlcv_df)
    _disable_history(mocker)  # sync 외부 호출 차단 — 이미 적재된 store 만 읽음
    _seed_store_aapl()

    spy_fund = mocker.spy(fundamentals, "fetch_fundamentals")
    spy_edgar = mocker.spy(edgar_client, "fetch_edgar_cached")
    spy_dart = mocker.spy(dart_client, "fetch_dart_cached")

    # write_portfolio_sheet 인자로 전달된 results 의 res.fundamentals 를 캡처.
    spy_write = mocker.spy(
        __import__("stocksig.main_run", fromlist=["write_portfolio_sheet"]),
        "write_portfolio_sheet",
    )

    tickers = tmp_tickers_file("AAPL\n")
    env = tmp_env_file(
        "EDGAR_USER_AGENT_EMAIL=test@example.com\nOPENDART_API_KEY=test-key\n"
    )

    output_path = run(tickers, env, tmp_path / "output")

    assert output_path.exists()
    # 외부 펀더멘털 fetch 호출 0 (compute_matrix 는 SQLite SELECT 만).
    assert spy_fund.call_count == 0
    assert spy_edgar.call_count == 0
    assert spy_dart.call_count == 0

    # write_portfolio_sheet 가 받은 results 의 AAPL fundamentals 가 store 매트릭스에서 왔다.
    assert spy_write.call_count == 1
    _wb, _formats, results, _failures, _input_order = spy_write.call_args.args[:5]
    by_sym = {r.spec.symbol: r for r in results}
    assert "AAPL" in by_sym
    fund = by_sym["AAPL"].fundamentals
    assert fund is not None
    # store 적재 종목 → 4셀 중 가격 무관(완성) GPM/OPM value 존재 (단일 원천 읽기 성공).
    assert fund.gpm.value is not None
    assert fund.opm.value is not None
    # provenance 라벨이 최신 분기(2026Q1)·소스(EDGAR)로 합성됨 (store 경유 단언).
    assert "2026Q1" in (fund.gpm.note or "")
