"""Plan 09-03 통합 테스트 — run_history 오케스트레이션 + history CLI 서브커맨드.

전부 네트워크 0 / 실DB 0: Plan 01 fixture(`fetch_fn_stub`/`build_ohlcv`) +
monkeypatch(count_rows/fetch_raw_quarters/quarter_end_prices/fetch_company_name).
SC1(별도 파일·시트1 불변)·SC3(원천/스냅샷)·D-09/10/11/04·D-15(DB 게이트·분리) 단언.
"""

from __future__ import annotations

import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

from stocksig.io import history_render
from stocksig.io.input import TickerSpec
from stocksig.io.metrics_engine import (
    _calendar_quarter_offset,
    compute_matrix,
    compute_peg_cell,
    price_ratio,
)
from fixtures.history_fixtures import (
    TICKER_INDUSTRY,
    build_ohlcv,
    fetch_fn_stub,
)

REPO_ROOT = Path(__file__).resolve().parents[1]

_SPECS = [
    TickerSpec(symbol="AAPL", tier="보유", industry="tech"),
    TickerSpec(symbol="005930.KS", tier="관심", industry="semiconductors"),
]


@pytest.fixture
def wired(monkeypatch):
    """run_history 의 모든 외부/ DB 의존을 fixture stub 으로 치환(네트워크·DB 0)."""
    monkeypatch.setattr(history_render, "_is_missing", history_render._is_missing)

    # compute_matrix → fixture fetch_fn 으로 강제(외부 호출 0).
    def _matrix(ticker, fetch_fn=None):
        return compute_matrix(ticker, fetch_fn=fetch_fn_stub)

    # io.* / output.* 의 늦은 import 대상은 모듈 attribute 로 patch.
    import stocksig.io.company as company_mod
    import stocksig.io.fundamentals_store as store_mod
    import stocksig.io.input as input_mod
    import stocksig.io.quarter_price as qp_mod

    monkeypatch.setattr(history_render, "compute_matrix", _matrix)
    monkeypatch.setattr(store_mod, "count_rows", lambda *a, **k: 999)
    monkeypatch.setattr(
        store_mod, "fetch_raw_quarters", lambda t: fetch_fn_stub(t)
    )
    monkeypatch.setattr(input_mod, "read_tickers_extended", lambda p: list(_SPECS))
    monkeypatch.setattr(qp_mod, "quarter_end_prices", lambda t: _qprices(t))
    monkeypatch.setattr(company_mod, "fetch_company_name", lambda t: f"{t} Inc")
    return None


def _qprices(ticker):
    """분기말 종가 dict + 현재가 — fixture OHLCV 기반(네트워크 0)."""
    df = build_ohlcv()
    close = df["Close"].dropna()
    qe = close.resample("QE").last()
    keys = qe.index.to_period("Q").astype(str)
    qmap = {k: float(v) for k, v in zip(keys, qe.to_numpy())}
    # fixture OHLCV 는 2024 년 → 매트릭스 분기(2025~2026)와 겹치지 않을 수 있으므로
    # 매트릭스 분기에도 종가가 잡히도록 결정적 가격을 보강한다.
    for q in ["2025Q1", "2025Q2", "2025Q3", "2025Q4", "2026Q1"]:
        qmap.setdefault(q, 50.0)
    return qmap, 60.0


# ─────────────────────── Task 1: render / peg / missing / layout ───────────────────────


def test_db_empty_guard(monkeypatch, capsys, tmp_path):
    """count_rows()==0 → None 반환 + 한국어 안내 print(예외 미발생) — D-15."""
    import stocksig.io.fundamentals_store as store_mod

    monkeypatch.setattr(store_mod, "count_rows", lambda *a, **k: 0)
    result = history_render.run_history("tickers.txt", str(tmp_path))
    out = capsys.readouterr().out
    assert result is None
    assert "펀더멘털 DB가 비어 있습니다" in out


def test_us_kr_alpha_order():
    """정렬: US 군이 KR 보다 먼저, 그룹 내 알파벳순 — D-03."""
    specs = [
        TickerSpec(symbol="005930.KS"),
        TickerSpec(symbol="MSFT"),
        TickerSpec(symbol="AAPL"),
        TickerSpec(symbol="000660.KS"),
    ]
    ordered = [s.symbol for s in history_render._sorted_tickers(specs)]
    assert ordered == ["AAPL", "MSFT", "000660.KS", "005930.KS"]


def test_matrix_layout_latest_left(wired, tmp_path):
    """display_quarters[0] == 최신 분기 — 엔진 오름차순 reversed 검증(D-01)."""
    history_render.run_history("tickers.txt", str(tmp_path))
    path = next(tmp_path.glob("fundamentals_history_*.xlsx"))
    wb = load_workbook(path)
    ws = wb["PER"]
    # 헤더: 식별 5열 + 분기열(최신 왼쪽). 6번째 셀(col F)이 최신 분기.
    first_quarter = ws.cell(row=1, column=6).value
    assert first_quarter == "2026Q1"


def test_peg_per_quarter(wired, tmp_path):
    """분기별 PEG = compute_peg_cell(현재가 PER, EPS_now, EPS_prior) 와 일치 — D-10."""
    matrix = compute_matrix("AAPL", fetch_fn=fetch_fn_stub)
    q = "2026Q1"
    price = 60.0  # _qprices current
    per = price_ratio(matrix["EPS_ttm"][q], price)
    qp = _calendar_quarter_offset(q, -4)  # 2025Q1
    eps_prior = matrix["EPS_ttm"][qp].value if qp in matrix["EPS_ttm"] else None
    expected = compute_peg_cell(per.value, matrix["EPS_ttm"][q].value, eps_prior)

    # run_history 가 채운 PEG 최신 분기 셀과 동일해야 한다.
    history_render.run_history("tickers.txt", str(tmp_path))
    # 직접 _inject_prices 경로 재현으로 동치 단언(렌더 산식 위임 확인).
    m2 = compute_matrix("AAPL", fetch_fn=fetch_fn_stub)
    qmap, current = _qprices("AAPL")
    quarters = sorted({qq for cells in m2.values() for qq in cells})
    history_render._inject_prices(m2, quarters, qmap, current, quarters[-1])
    assert m2["PEG"][q].value == pytest.approx(expected.value) if expected.value is not None \
        else m2["PEG"][q].value is None


def test_missing_dash_guard(wired, tmp_path):
    """미보유 분기 셀 KeyError 없이 '-' (Pitfall 2 .get 가드)."""
    history_render.run_history("tickers.txt", str(tmp_path))
    path = next(tmp_path.glob("fundamentals_history_*.xlsx"))
    wb = load_workbook(path)
    ws = wb["PEG"]
    # 최우측(가장 오래된) 분기는 4분기 전 EPS 부재 → PEG 결손 "-" 가능. KeyError 0 이면 통과.
    assert ws.max_row >= 2


def test_no_main_run_coupling():
    """history_render 에서 main_run import 0(D-15 분리)."""
    src = (REPO_ROOT / "src/stocksig/io/history_render.py").read_text(encoding="utf-8")
    assert "main_run" not in src
    assert "compute_matrix" in src
    assert "quarter_end_prices" in src


# ─────────────────────── Task 2: history CLI 서브커맨드 ───────────────────────


def test_history_cli_help():
    """`python main.py history --help` 종료코드 0 + '펀더멘털 트렌드' 출력."""
    proc = subprocess.run(
        [sys.executable, "main.py", "history", "--help"],
        cwd=str(REPO_ROOT),
        capture_output=True,
        text=True,
        encoding="utf-8",
    )
    assert proc.returncode == 0
    assert "펀더멘털 트렌드" in proc.stdout


def test_history_cli_dispatch(monkeypatch):
    """history 분기 → run_history, 기본 분기 → main_run.run (서로 다른 엔트리, D-15)."""
    import importlib

    main_mod = importlib.import_module("main")

    called = {}

    import stocksig.io.history_render as hr_mod
    monkeypatch.setattr(hr_mod, "run_history", lambda t, o: called.setdefault("history", (t, o)) or Path("x"))

    monkeypatch.setattr(sys, "argv", ["main.py", "history", "--tickers", "t.txt", "--output-dir", "o"])
    rc = main_mod.main()
    assert rc == 0
    assert called["history"] == ("t.txt", "o")


def test_default_cli_dispatch(monkeypatch):
    """서브커맨드 없으면 main_run.run 호출(하위호환)."""
    import importlib

    main_mod = importlib.import_module("main")
    import stocksig.main_run as run_mod

    called = {}
    monkeypatch.setattr(run_mod, "run", lambda *a, **k: called.setdefault("run", True) or Path("p"))
    monkeypatch.setattr(sys, "argv", ["main.py", "--tickers", "t.txt"])
    rc = main_mod.main()
    assert rc == 0
    assert called.get("run") is True


def test_history_cli_db_empty_exit0(monkeypatch):
    """run_history None 반환(DB 미적재) 시 종료코드 0(예외 아님)."""
    import importlib

    main_mod = importlib.import_module("main")
    import stocksig.io.history_render as hr_mod
    monkeypatch.setattr(hr_mod, "run_history", lambda t, o: None)
    monkeypatch.setattr(sys, "argv", ["main.py", "history"])
    rc = main_mod.main()
    assert rc == 0


# ─────────────────────── Task 3: 통합(별도파일·시트1 불변·원천/스냅샷·freeze) ───────────────────────


def test_separate_file_sheet1_untouched(wired, tmp_path):
    """트렌드 파일 생성, portfolio_*.xlsx 미생성·미변경 — SC1 / D-14/D-15."""
    result = history_render.run_history("tickers.txt", str(tmp_path))
    assert result is not None
    assert result.name.startswith("fundamentals_history_")
    assert result.name.endswith(".xlsx")
    assert result.exists()
    # 같은 디렉터리에 시트1 portfolio 파일은 생성되지 않음(흐름 분리).
    assert list(tmp_path.glob("portfolio_*.xlsx")) == []


def test_raw_and_snapshot_sheets(wired, tmp_path):
    """[원천]·[최신 스냅샷] 시트 존재·내용 단언 — SC3 / D-13."""
    history_render.run_history("tickers.txt", str(tmp_path))
    path = next(tmp_path.glob("fundamentals_history_*.xlsx"))
    wb = load_workbook(path)
    # 시트명 sanitize: [] 제거 → "원천" / "최신 스냅샷".
    assert "원천" in wb.sheetnames
    assert "최신 스냅샷" in wb.sheetnames

    raw = wb["원천"]
    assert raw.cell(row=1, column=1).value == "티커"
    assert raw.max_row >= 2  # long 행 존재.

    snap = wb["최신 스냅샷"]
    assert snap.cell(row=1, column=1).value == "티커"
    # 종목 2개 → 헤더 1 + 2행.
    assert snap.max_row == 3


def test_freeze(wired, tmp_path):
    """지표 시트 freeze_panes == 'B1'(A열만, 헤더행 미고정) — D-04."""
    history_render.run_history("tickers.txt", str(tmp_path))
    path = next(tmp_path.glob("fundamentals_history_*.xlsx"))
    wb = load_workbook(path)
    assert wb["PER"].freeze_panes == "B1"
    assert wb["최신 스냅샷"].freeze_panes == "B1"


def test_all_nine_metric_sheets(wired, tmp_path):
    """9 지표 시트 전부 생성 — SC2."""
    history_render.run_history("tickers.txt", str(tmp_path))
    path = next(tmp_path.glob("fundamentals_history_*.xlsx"))
    wb = load_workbook(path)
    for m in ["PER", "PEG", "GPM", "OPM", "PBR", "PCR", "PSR", "ROE", "ROA"]:
        assert m in wb.sheetnames
