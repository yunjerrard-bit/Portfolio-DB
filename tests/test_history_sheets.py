"""트렌드 히스토리 시트 writer read-back 단언 (Plan 09-02) — 네트워크 0.

openpyxl 정적 read-back 으로 검증:
  - matrix: 식별 5열 + 분기열(최신 왼쪽), 결손 "-"+코멘트(D-11), 상대색 fill(D-05/06),
    YoY 글리프(D-08), 비율 지표·ROE/ROA 퍼센트 표기(WARNING-2), 헤더행+A열 freeze(D-04).
  - raw: [원천] long 시트 헤더 + 7-tuple 행, 헤더행만 freeze(A2).
  - snapshot: [최신 스냅샷] 식별 5열 + 9지표, 최신값 재사용, PEG 결손 "-", ROE/ROA 퍼센트.

전부 합성 MetricCell / fixture — yfinance·네트워크 호출 0.
"""

from __future__ import annotations

import openpyxl

from stocksig.io.fundamentals import MetricCell
from stocksig.output.history_workbook import make_history_workbook
from stocksig.output.sheet_metric_matrix import write_metric_sheet
from stocksig.output.sheet_raw import write_raw_sheet
from stocksig.output.sheet_snapshot import write_snapshot_sheet


def _cell(value, source="EDGAR", note=None):
    return MetricCell(value=value, source=source, note=note)


# ---------------- matrix --------------------------------------------------


def _write_matrix(path, metric, ticker_rows, display_quarters, peer_lookup, prior_lookup):
    wb, formats = make_history_workbook(path)
    ws = wb.add_worksheet(metric)
    write_metric_sheet(
        wb, ws, metric, ticker_rows, display_quarters, formats, peer_lookup, prior_lookup
    )
    wb.close()


def test_matrix_headers_and_freeze(tmp_path):
    """식별 5열 헤더 + 최신 분기 왼쪽(D-01) + 헤더행+A열 freeze(D-04)."""
    path = str(tmp_path / "h.xlsx")
    display_quarters = ["2026Q1", "2025Q4", "2025Q3", "2025Q2", "2025Q1"]
    rows = [
        {
            "ticker": "AAPL",
            "company": "Apple Inc",
            "market": "US",
            "tier": "A",
            "industry": "tech",
            "cells": {q: _cell(10.0 + i) for i, q in enumerate(display_quarters)},
        },
    ]
    _write_matrix(
        path, "PER", rows, display_quarters,
        peer_lookup=lambda m, q, ind: [], prior_lookup=lambda m, t, q: None,
    )
    wb = openpyxl.load_workbook(path)
    ws = wb["PER"]
    assert ws["A1"].value == "티커"
    assert ws["B1"].value == "기업명"
    assert ws["E1"].value == "산업"
    assert ws["F1"].value == display_quarters[0] == "2026Q1"  # 최신 왼쪽
    assert ws.freeze_panes == "B2"  # 헤더행+A열 고정, B2 부터 스크롤


def test_matrix_missing_cell_dash_and_comment(tmp_path):
    """결손 셀 = "-" + 사유 코멘트(D-11, 0/빈칸 금지)."""
    path = str(tmp_path / "h.xlsx")
    display_quarters = ["2026Q1", "2025Q4"]
    rows = [
        {
            "ticker": "AAPL", "company": "Apple", "market": "US",
            "tier": "A", "industry": "tech",
            "cells": {
                "2026Q1": _cell(None, source=None, note="조회 실패: EPS ≤ 0"),
                # 2025Q4 미보유 → .get None
            },
        },
    ]
    _write_matrix(
        path, "PER", rows, display_quarters,
        peer_lookup=lambda m, q, ind: [], prior_lookup=lambda m, t, q: None,
    )
    wb = openpyxl.load_workbook(path)
    ws = wb["PER"]
    assert ws["F2"].value == "-"  # 결손
    assert ws["F2"].comment is not None
    assert "EPS" in ws["F2"].comment.text
    assert ws["G2"].value == "-"  # 분기 미보유도 "-"
    assert ws["G2"].comment is not None


def test_matrix_relative_color_and_sample_gate(tmp_path):
    """표본<3 무색(D-07), 표본≥3 LOWER_IS_BETTER 최저값=초록 fill(D-05/06)."""
    path = str(tmp_path / "h.xlsx")
    q = "2026Q1"
    display_quarters = [q]
    rows = [
        {"ticker": "LOW", "company": "L", "market": "US", "tier": "A",
         "industry": "tech", "cells": {q: _cell(5.0)}},
        {"ticker": "MID", "company": "M", "market": "US", "tier": "A",
         "industry": "tech", "cells": {q: _cell(15.0)}},
        {"ticker": "HIGH", "company": "H", "market": "US", "tier": "A",
         "industry": "tech", "cells": {q: _cell(25.0)}},
        {"ticker": "SOLO", "company": "S", "market": "US", "tier": "A",
         "industry": "lonely", "cells": {q: _cell(99.0)}},
    ]
    peer_by_ind = {"tech": [5.0, 15.0, 25.0], "lonely": [99.0]}

    def peer_lookup(metric, quarter, industry):
        return peer_by_ind.get(industry, [])

    _write_matrix(
        path, "PER", rows, display_quarters,
        peer_lookup=peer_lookup, prior_lookup=lambda m, t, q: None,
    )
    wb = openpyxl.load_workbook(path)
    ws = wb["PER"]
    # LOW (PER 최저, LOWER_IS_BETTER) → 초록 fill C8E6C9
    low_fill = ws["F2"].fill.fgColor.rgb
    assert str(low_fill).endswith("C8E6C9")
    # SOLO 산업 표본<3 → 무색(fill 미설정)
    solo_fill = ws["F5"].fill.fgColor.rgb
    assert not str(solo_fill).endswith("C8E6C9")
    assert not str(solo_fill).endswith("FFCDD2")


def test_matrix_yoy_glyph_and_prior_missing(tmp_path):
    """YoY 증가 셀 ▲(D-08), 전년 결손 셀 글리프 없음."""
    path = str(tmp_path / "h.xlsx")
    q = "2026Q1"
    display_quarters = [q]
    rows = [
        {"ticker": "UP", "company": "U", "market": "US", "tier": "A",
         "industry": "tech", "cells": {q: _cell(20.0)}},
        {"ticker": "NOPRIOR", "company": "N", "market": "US", "tier": "A",
         "industry": "tech", "cells": {q: _cell(20.0)}},
    ]

    def prior_lookup(metric, ticker, quarter):
        if ticker == "UP":
            return _cell(10.0)  # 전년 10 < 현재 20 → ▲
        return None  # 전년 결손

    _write_matrix(
        path, "PER", rows, display_quarters,
        peer_lookup=lambda m, q, ind: [], prior_lookup=prior_lookup,
    )
    wb = openpyxl.load_workbook(path)
    ws = wb["PER"]
    assert "▲" in ws["F2"].value
    assert "▲" not in ws["F3"].value and "▼" not in ws["F3"].value


def test_matrix_ratio_percent_vs_decimal(tmp_path):
    """비율 지표(GPM) 퍼센트 표기, 비-비율(PER) 소수 표기 — WARNING-2 시트1 정합."""
    q = "2026Q1"
    display_quarters = [q]
    # GPM 0.27 → "27.0%"
    gpm_rows = [{"ticker": "A", "company": "A", "market": "US", "tier": "A",
                 "industry": "tech", "cells": {q: _cell(0.27)}}]
    p1 = "x.xlsx"
    import tempfile
    import os
    d = tempfile.mkdtemp()
    p_gpm = os.path.join(d, "gpm.xlsx")
    _write_matrix(
        p_gpm, "GPM", gpm_rows, display_quarters,
        peer_lookup=lambda m, q, ind: [], prior_lookup=lambda m, t, q: None,
    )
    ws = openpyxl.load_workbook(p_gpm)["GPM"]
    assert "%" in ws["F2"].value and "27.0%" in ws["F2"].value

    per_rows = [{"ticker": "A", "company": "A", "market": "US", "tier": "A",
                 "industry": "tech", "cells": {q: _cell(15.5)}}]
    p_per = os.path.join(d, "per.xlsx")
    _write_matrix(
        p_per, "PER", per_rows, display_quarters,
        peer_lookup=lambda m, q, ind: [], prior_lookup=lambda m, t, q: None,
    )
    ws2 = openpyxl.load_workbook(p_per)["PER"]
    assert "%" not in ws2["F2"].value and "." in ws2["F2"].value


def test_matrix_roe_roa_percent(tmp_path):
    """ROE/ROA 는 registry is_ratio_0_1=False 지만 매트릭스에서 퍼센트(.1f%)로 표기.

    ROE 1.151 → "115.1%" (100% 초과 가능 — AAPL 실측). registry 플래그 불변.
    """
    q = "2026Q1"
    display_quarters = [q]
    import os
    import tempfile

    d = tempfile.mkdtemp()

    roe_rows = [{"ticker": "AAPL", "company": "A", "market": "US", "tier": "A",
                 "industry": "tech", "cells": {q: _cell(1.151)}}]
    p_roe = os.path.join(d, "roe.xlsx")
    _write_matrix(
        p_roe, "ROE", roe_rows, display_quarters,
        peer_lookup=lambda m, q, ind: [], prior_lookup=lambda m, t, q: None,
    )
    ws_roe = openpyxl.load_workbook(p_roe)["ROE"]
    assert "%" in ws_roe["F2"].value and "115.1%" in ws_roe["F2"].value

    roa_rows = [{"ticker": "AAPL", "company": "A", "market": "US", "tier": "A",
                 "industry": "tech", "cells": {q: _cell(0.09)}}]
    p_roa = os.path.join(d, "roa.xlsx")
    _write_matrix(
        p_roa, "ROA", roa_rows, display_quarters,
        peer_lookup=lambda m, q, ind: [], prior_lookup=lambda m, t, q: None,
    )
    ws_roa = openpyxl.load_workbook(p_roa)["ROA"]
    assert "%" in ws_roa["F2"].value and "9.0%" in ws_roa["F2"].value


# ---------------- raw -----------------------------------------------------


def test_raw_sheet_long_rows(tmp_path):
    """[원천] long 시트 — 헤더 + 7-tuple 행, 값 결손 "-" 일관."""
    path = str(tmp_path / "h.xlsx")
    wb, formats = make_history_workbook(path)
    # Excel 금지문자 [] → 시트명은 호출자(Plan 03)가 sanitize. 검증은 sanitize명으로.
    ws = wb.add_worksheet("원천")
    raw_by_ticker = {
        "AAPL": [
            ("2026Q1", "EDGAR", "revenue", 1000.0, "duration", None, "USD"),
            ("2026Q1", "EDGAR", "net_income", None, "duration", None, "USD"),
        ],
    }
    write_raw_sheet(ws, raw_by_ticker, formats, sorted_tickers=["AAPL"])
    wb.close()

    wb2 = openpyxl.load_workbook(path)
    ws2 = wb2["원천"]
    # 헤더 한국어
    headers = [ws2.cell(row=1, column=c).value for c in range(1, 9)]
    assert "티커" in headers and "소스" in headers and "분기" in headers
    # 데이터 행: 티커 + raw 값
    assert ws2.cell(row=2, column=1).value == "AAPL"
    # 결손 value → "-"
    row3 = [ws2.cell(row=3, column=c).value for c in range(1, 9)]
    assert "-" in row3
    # 원천 시트: 헤더행만 freeze(A2, 키 컬럼 미고정) — D-04.
    assert ws2.freeze_panes == "A2"


# ---------------- snapshot -------------------------------------------------


def test_snapshot_sheet_one_row_per_ticker(tmp_path):
    """[최신 스냅샷] 식별 5열 + 9지표, 최신값 재사용, PEG 결손 "-"."""
    path = str(tmp_path / "h.xlsx")
    wb, formats = make_history_workbook(path)
    ws = wb.add_worksheet("최신 스냅샷")
    snapshot_rows = [
        {
            "ticker": "AAPL", "company": "Apple", "market": "US",
            "tier": "A", "industry": "tech",
            "metrics": {
                "PER": _cell(15.5),
                "PEG": _cell(None, source=None, note="성장률 ≤ 0"),
                "GPM": _cell(0.27),
                "OPM": _cell(0.20),
                "PBR": _cell(5.0),
                "PCR": _cell(10.0),
                "PSR": _cell(3.0),
                "ROE": _cell(0.18),
                "ROA": _cell(0.09),
            },
        },
    ]
    write_snapshot_sheet(ws, snapshot_rows, formats)
    wb.close()

    wb2 = openpyxl.load_workbook(path)
    ws2 = wb2["최신 스냅샷"]
    assert ws2["A1"].value == "티커"
    headers = [ws2.cell(row=1, column=c).value for c in range(1, 15)]
    for m in ["PER", "PEG", "GPM", "OPM", "PBR", "PCR", "PSR", "ROE", "ROA"]:
        assert m in headers
    # 최신값 재사용
    assert ws2["A2"].value == "AAPL"
    # PER 셀에 15.5 표기(소수)
    per_col = headers.index("PER") + 1
    assert "15.5" in str(ws2.cell(row=2, column=per_col).value)
    # PEG 결손 → "-"
    peg_col = headers.index("PEG") + 1
    assert ws2.cell(row=2, column=peg_col).value == "-"
    assert ws2.cell(row=2, column=peg_col).comment is not None
    # ROE 0.18 → "18.0%", ROA 0.09 → "9.0%" (퍼센트 표기, registry 불변).
    roe_col = headers.index("ROE") + 1
    roe_val = str(ws2.cell(row=2, column=roe_col).value)
    assert "%" in roe_val and "18.0%" in roe_val
    roa_col = headers.index("ROA") + 1
    roa_val = str(ws2.cell(row=2, column=roa_col).value)
    assert "%" in roa_val and "9.0%" in roa_val
    # 스냅샷 헤더행+A열 freeze(B2) — D-04.
    assert ws2.freeze_panes == "B2"
