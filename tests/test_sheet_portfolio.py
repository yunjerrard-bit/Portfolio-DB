"""Phase 2 Wave 3 (02-03): 시트1 (포트폴리오 통합) writer 테스트.

D-02 (단일 색 결정 경로), D-03 (실패 행), D-08 (15 컬럼), PORT-02/03/04/06/07/08, TECH-07.

전부 합성 DataFrame — yfinance 호출 없음. 읽기는 openpyxl로 정적 검증.
"""

from __future__ import annotations

import re
from datetime import datetime

import openpyxl
import pandas as pd
import pytest

from stocksig.io.fundamentals import FundamentalsResult, MetricCell
from stocksig.io.input import TickerSpec
from stocksig.output.sheet_portfolio import (
    PORTFOLIO_COLUMNS,
    write_portfolio_sheet,
)
from stocksig.output.writer import make_workbook
from stocksig.runner import TickerFailure, TickerResult


# ---------------- fixtures / helpers --------------------------------------


def _make_enriched_df(
    n_rows: int = 10,
    close: float = 100.0,
    close_pct: float = 0.005,
    diff11: float = 0.0,
    diff11_med: float = 0.0,
    diff11_std: float = 0.01,
    volume: float = 1_000_000.0,
    vol_pct: float = 0.05,
    vol_pct_med: float = 0.0,
    vol_pct_std: float = 0.1,
    stoch_k: float = 50.0,
    rsi: float = 50.0,
    impulse_daily: str = "녹색",
    impulse_weekly: str = "청색",
) -> pd.DataFrame:
    """10개 비즈니스 데이 ascending index — Phase 1 pipeline 출력 형태와 동일."""
    idx = pd.bdate_range("2026-05-12", periods=n_rows)
    row = {
        "Close": close,
        "Close_pct_change": close_pct,
        "Close_pct_change_median": 0.001,
        "Close_pct_change_std": 0.01,
        "EMA_Close_11": close,
        "EMA_Close_22": close,
        "EMA_Close_96": close,
        "EMA_Close_192": close,
        "DIFF_Close_11": diff11,
        "DIFF_Close_11_median": diff11_med,
        "DIFF_Close_11_std": diff11_std,
        "DIFF_Close_22": 0.0,
        "DIFF_Close_22_median": 0.0,
        "DIFF_Close_22_std": 0.01,
        "DIFF_Close_96": 0.0,
        "DIFF_Close_96_median": 0.0,
        "DIFF_Close_96_std": 0.01,
        "DIFF_Close_192": 0.0,
        "DIFF_Close_192_median": 0.0,
        "DIFF_Close_192_std": 0.01,
        "Volume": volume,
        "Volume_pct_change": vol_pct,
        "Volume_pct_change_median": vol_pct_med,
        "Volume_pct_change_std": vol_pct_std,
        "Stoch_%K": stoch_k,
        "RSI": rsi,
        "Impulse_daily": impulse_daily,
        "Impulse_weekly": impulse_weekly,
    }
    return pd.DataFrame([row] * n_rows, index=idx)


def _spec(symbol: str = "AAPL", tier: str = "1", industry: str = "Technology") -> TickerSpec:
    return TickerSpec(symbol=symbol, tier=tier, industry=industry)


def _result(spec: TickerSpec, market: str = "US", **df_kwargs) -> TickerResult:
    return TickerResult(spec=spec, enriched_df=_make_enriched_df(**df_kwargs), market=market)


def _open(path):
    return openpyxl.load_workbook(str(path))


# ---------------- tests ----------------------------------------------------


def test_column_count_is_21():
    # 03-05: 기존 17열 + PER/PEG/GPM/OPM 4열 = 21열 (D-01, index 17~20).
    assert len(PORTFOLIO_COLUMNS) == 21
    assert PORTFOLIO_COLUMNS[-4:] == ["PER", "PEG", "GPM", "OPM"]


def test_column_order(tmp_path):
    path = tmp_path / "out.xlsx"
    wb, formats = make_workbook(path)
    write_portfolio_sheet(wb, formats, [_result(_spec())], [], ["AAPL"], now=datetime(2026, 5, 26, 12, 0, 0))
    wb.close()

    ws = _open(path)["시트1"]
    for col_idx, name in enumerate(PORTFOLIO_COLUMNS, start=1):
        assert ws.cell(row=5, column=col_idx).value == name, (
            f"header col {col_idx} mismatch: expected {name}, got {ws.cell(5, col_idx).value}"
        )


def test_timestamp_row1(tmp_path):
    path = tmp_path / "out.xlsx"
    wb, formats = make_workbook(path)
    write_portfolio_sheet(
        wb, formats, [_result(_spec())], [], ["AAPL"], now=datetime(2026, 5, 26, 12, 0, 0)
    )
    wb.close()

    ws = _open(path)["시트1"]
    assert re.match(r"^실행 시각: \d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}$", ws["A1"].value)


def test_one_row_per_ticker_in_input_order(tmp_path):
    path = tmp_path / "out.xlsx"
    wb, formats = make_workbook(path)
    specs = [_spec("AAPL"), _spec("MSFT"), _spec("GOOG")]
    # Pass results in reversed order (simulate as_completed race)
    results = [_result(s) for s in reversed(specs)]
    write_portfolio_sheet(wb, formats, results, [], ["AAPL", "MSFT", "GOOG"], now=datetime(2026, 5, 26))
    wb.close()

    ws = _open(path)["시트1"]
    assert ws.cell(row=6, column=1).value == "AAPL"
    assert ws.cell(row=7, column=1).value == "MSFT"
    assert ws.cell(row=8, column=1).value == "GOOG"


def test_columns_ticker_market_close_change(tmp_path):
    path = tmp_path / "out.xlsx"
    wb, formats = make_workbook(path)
    write_portfolio_sheet(
        wb, formats, [_result(_spec(), market="US")], [], ["AAPL"], now=datetime(2026, 5, 26)
    )
    wb.close()

    ws = _open(path)["시트1"]
    assert ws.cell(row=6, column=1).value == "AAPL"
    assert ws.cell(row=6, column=2).value == "US"
    assert ws.cell(row=6, column=5).value == pytest.approx(100.0)
    assert ws.cell(row=6, column=6).value == pytest.approx(0.005)


def test_ticker_hyperlink_us_and_kr(tmp_path):
    path = tmp_path / "out.xlsx"
    wb, formats = make_workbook(path)
    results = [
        _result(_spec("AAPL"), market="US"),
        _result(_spec("005930.KS"), market="KR"),
    ]
    write_portfolio_sheet(wb, formats, results, [], ["AAPL", "005930.KS"], now=datetime(2026, 5, 26))
    wb.close()

    ws = _open(path)["시트1"]
    a6 = ws.cell(row=6, column=1)
    a7 = ws.cell(row=7, column=1)
    assert a6.hyperlink is not None
    # openpyxl exposes internal link via .location, target may be None
    assert a6.hyperlink.location.endswith("'AAPL'!A1")
    assert a7.hyperlink.location.endswith("'005930.KS'!A1")


def test_diff_ema_4cells_color_soft_red(tmp_path):
    """DIFF_Close_11 deviation = median + 1.5*std → SOFT_RED. G6 font.color == RED_800."""
    path = tmp_path / "out.xlsx"
    wb, formats = make_workbook(path)
    # 1.5σ above median => SOFT_RED
    res = _result(_spec(), diff11=0.0 + 1.5 * 0.01, diff11_med=0.0, diff11_std=0.01)
    write_portfolio_sheet(wb, formats, [res], [], ["AAPL"], now=datetime(2026, 5, 26))
    wb.close()

    ws = _open(path)["시트1"]
    g6 = ws.cell(row=6, column=7)
    # font.color may be ARGB "FFC62828" or theme — accept either RGB end-match
    color_val = g6.font.color.value if g6.font.color else ""
    assert isinstance(color_val, str)
    assert color_val.upper().endswith("C62828"), f"expected ...C62828 (RED_800), got {color_val!r}"


def test_diff_color_matches_per_ticker_logic():
    """D-02 single source of truth: portfolio DIFF color decision uses the
    same `decide_sigma_bucket` function as the per-ticker sheet. Cross-verify
    enum identity (planner allowed simplification: comparing SigmaBucket enum)."""
    from stocksig.compute.color_rules import decide_sigma_bucket
    from stocksig.output import sheet_portfolio  # noqa: F401

    # Same inputs must produce identical bucket — by definition (same fn).
    val, med, std = 0.025, 0.0, 0.01
    bucket_a = decide_sigma_bucket(val, med, std)
    # sheet_portfolio module must import & call this exact function.
    src = open(sheet_portfolio.__file__, encoding="utf-8").read()
    assert "decide_sigma_bucket" in src, (
        "sheet_portfolio.py must import decide_sigma_bucket from compute.color_rules"
    )
    # And the enum is the canonical SigmaBucket.HARD_RED for 2.5σ above:
    from stocksig.compute.color_rules import SigmaBucket
    assert bucket_a == SigmaBucket.HARD_RED


def test_volume_color(tmp_path):
    """K6 (거래량) uses decide_sigma_bucket(vol_pct, med, std)."""
    path = tmp_path / "out.xlsx"
    wb, formats = make_workbook(path)
    # vol_pct = 0.25, med=0, std=0.1 → deviation 2.5σ → HARD_RED
    res = _result(_spec(), vol_pct=0.25, vol_pct_med=0.0, vol_pct_std=0.1)
    write_portfolio_sheet(wb, formats, [res], [], ["AAPL"], now=datetime(2026, 5, 26))
    wb.close()

    ws = _open(path)["시트1"]
    k6 = ws.cell(row=6, column=11)
    color_val = k6.font.color.value if k6.font.color else ""
    # HARD_RED font is RED_900 = #B71C1C
    assert color_val.upper().endswith("B71C1C"), f"expected ...B71C1C (RED_900 HARD_RED), got {color_val!r}"


def test_stoch_rsi_color(tmp_path):
    """L6 (%K=15) SOFT_GREEN; M6 (RSI=75) SOFT_RED."""
    path = tmp_path / "out.xlsx"
    wb, formats = make_workbook(path)
    res = _result(_spec(), stoch_k=15.0, rsi=75.0)
    write_portfolio_sheet(wb, formats, [res], [], ["AAPL"], now=datetime(2026, 5, 26))
    wb.close()

    ws = _open(path)["시트1"]
    l6 = ws.cell(row=6, column=12)
    m6 = ws.cell(row=6, column=13)
    l_color = l6.font.color.value if l6.font.color else ""
    m_color = m6.font.color.value if m6.font.color else ""
    assert l_color.upper().endswith("2E7D32"), f"%K=15 expected GREEN_800, got {l_color!r}"
    assert m_color.upper().endswith("C62828"), f"RSI=75 expected RED_800, got {m_color!r}"


def test_impulse_cells(tmp_path):
    """P6=녹색 (GREEN_800), Q6=청색 (BLUE_800)."""
    path = tmp_path / "out.xlsx"
    wb, formats = make_workbook(path)
    res = _result(_spec(), impulse_daily="녹색", impulse_weekly="청색")
    write_portfolio_sheet(wb, formats, [res], [], ["AAPL"], now=datetime(2026, 5, 26))
    wb.close()

    ws = _open(path)["시트1"]
    n6 = ws.cell(row=6, column=16)
    o6 = ws.cell(row=6, column=17)
    assert n6.value == "녹색"
    assert o6.value == "청색"
    n_color = n6.font.color.value if n6.font.color else ""
    o_color = o6.font.color.value if o6.font.color else ""
    assert n_color.upper().endswith("2E7D32")  # GREEN_800
    assert o_color.upper().endswith("1565C0")  # BLUE_800


def test_tier_industry_columns(tmp_path):
    path = tmp_path / "out.xlsx"
    wb, formats = make_workbook(path)
    res = _result(_spec(tier="1", industry="Technology"))
    write_portfolio_sheet(wb, formats, [res], [], ["AAPL"], now=datetime(2026, 5, 26))
    wb.close()

    ws = _open(path)["시트1"]
    assert ws.cell(row=6, column=3).value == "1"
    assert ws.cell(row=6, column=4).value == "Technology"


def test_failed_row_in_sheet1(tmp_path):
    """D-03: failed ticker → trailing row, A=ticker B='?' middle empty, last col 실패: <reason>."""
    path = tmp_path / "out.xlsx"
    wb, formats = make_workbook(path)
    failures = [TickerFailure(spec=_spec("XYZ"), reason="부분 데이터: 100 거래일")]
    write_portfolio_sheet(wb, formats, [], failures, ["XYZ"], now=datetime(2026, 5, 26))
    wb.close()

    ws = _open(path)["시트1"]
    # No successes → first failure row is row 6.
    assert ws.cell(row=6, column=1).value == "XYZ"
    assert ws.cell(row=6, column=2).value == "?"
    # Middle cells empty:
    assert ws.cell(row=6, column=5).value in (None, "")
    # Last column (17) = 실패: <reason>
    last = ws.cell(row=6, column=17).value
    assert last is not None and last.startswith("실패: ")
    assert "부분 데이터" in last
    # Italic + bg fill present (failed_row_marker)
    assert ws.cell(row=6, column=17).font.italic is True


def test_input_order_preserved_with_mixed_success_fail(tmp_path):
    """Successes in input order rows 6..; failures appended in input order after."""
    path = tmp_path / "out.xlsx"
    wb, formats = make_workbook(path)
    results = [
        _result(_spec("MSFT")),
        _result(_spec("AAPL")),
    ]
    failures = [TickerFailure(spec=_spec("BAD"), reason="네트워크 오류")]
    write_portfolio_sheet(
        wb, formats, results, failures, ["AAPL", "MSFT", "BAD"], now=datetime(2026, 5, 26)
    )
    wb.close()

    ws = _open(path)["시트1"]
    # AAPL first (input idx 0), MSFT next, then BAD failure.
    assert ws.cell(row=6, column=1).value == "AAPL"
    assert ws.cell(row=7, column=1).value == "MSFT"
    assert ws.cell(row=8, column=1).value == "BAD"
    assert ws.cell(row=8, column=2).value == "?"


# ---------------- 03-05: 펀더멘털 4열 (PORT-05 / FUND-05 / D-05) ----------


def _fund(per=None, peg=None, gpm=None, opm=None) -> FundamentalsResult:
    """MetricCell 4종 헬퍼 — 각 인자는 (value, source, note) 튜플 또는 None."""
    def cell(t):
        if t is None:
            return MetricCell(value=None, source=None, note=None)
        v, s, n = t
        return MetricCell(value=v, source=s, note=n)

    return FundamentalsResult(
        per=cell(per), peg=cell(peg), gpm=cell(gpm), opm=cell(opm)
    )


def test_fund_cols(tmp_path):
    """PORT-05: col 18~21(PER/PEG/GPM/OPM) 값 + 출처 셀 주석, 포맷 무손상."""
    path = tmp_path / "out.xlsx"
    wb, formats = make_workbook(path)
    fund = _fund(
        per=(15.5, "EDGAR", "EDGAR · 2026Q2"),
        peg=(1.25, "EDGAR", "EDGAR · 2026Q2"),
        gpm=(0.45, "EDGAR", "EDGAR · 2026Q2"),
        opm=(0.30, "EDGAR", "EDGAR · 2026Q2"),
    )
    res = TickerResult(
        spec=_spec(), enriched_df=_make_enriched_df(), market="US", fundamentals=fund
    )
    write_portfolio_sheet(wb, formats, [res], [], ["AAPL"], now=datetime(2026, 5, 26))
    wb.close()

    ws = _open(path)["시트1"]
    per_c = ws.cell(row=6, column=18)
    peg_c = ws.cell(row=6, column=19)
    gpm_c = ws.cell(row=6, column=20)
    opm_c = ws.cell(row=6, column=21)

    assert per_c.value == pytest.approx(15.5)
    assert peg_c.value == pytest.approx(1.25)
    assert gpm_c.value == pytest.approx(0.45)
    assert opm_c.value == pytest.approx(0.30)

    # FUND-05: 출처 셀 주석 (write_comment)
    assert per_c.comment is not None
    assert "EDGAR" in per_c.comment.text
    assert gpm_c.comment is not None and "EDGAR" in gpm_c.comment.text

    # num_format 무손상: PER/PEG = #,##0.00, GPM/OPM = 0.00%
    assert per_c.number_format == "#,##0.00"
    assert peg_c.number_format == "#,##0.00"
    assert gpm_c.number_format == "0.00%"
    assert opm_c.number_format == "0.00%"


def test_fund_missing_cell_blank_with_note(tmp_path):
    """D-05: 결손 셀은 빈 칸 + 한국어 사유 주석 (0/특수값 미사용)."""
    path = tmp_path / "out.xlsx"
    wb, formats = make_workbook(path)
    fund = _fund(
        per=(15.5, "EDGAR", "EDGAR · 2026Q2"),
        peg=None,  # 결손
        gpm=(0.45, "EDGAR", "EDGAR · 2026Q2"),
        opm=None,  # 결손
    )
    # 결손 셀 note 부여 (fetch 단이 사유를 채운 형태 모사)
    fund.peg.note = "조회 실패: EPS 성장률 ≤ 0"
    res = TickerResult(
        spec=_spec(), enriched_df=_make_enriched_df(), market="US", fundamentals=fund
    )
    write_portfolio_sheet(wb, formats, [res], [], ["AAPL"], now=datetime(2026, 5, 26))
    wb.close()

    ws = _open(path)["시트1"]
    peg_c = ws.cell(row=6, column=19)
    opm_c = ws.cell(row=6, column=21)
    # 값 없음 (0 금지)
    assert peg_c.value in (None, "")
    assert opm_c.value in (None, "")
    # 사유 주석 존재
    assert peg_c.comment is not None
    assert "EPS 성장률" in peg_c.comment.text


def test_fund_none_backward_compat(tmp_path):
    """하위호환: res.fundamentals=None 시 4셀 빈칸 + '펀더멘털 미수집' 주석, raise 없음."""
    path = tmp_path / "out.xlsx"
    wb, formats = make_workbook(path)
    res = TickerResult(
        spec=_spec(), enriched_df=_make_enriched_df(), market="US", fundamentals=None
    )
    write_portfolio_sheet(wb, formats, [res], [], ["AAPL"], now=datetime(2026, 5, 26))
    wb.close()

    ws = _open(path)["시트1"]
    for col in (18, 19, 20, 21):
        c = ws.cell(row=6, column=col)
        assert c.value in (None, "")
        assert c.comment is not None
        assert "펀더멘털 미수집" in c.comment.text


def test_fund_failure_row_no_fund_cells(tmp_path):
    """D-06: 실패 행은 펀더멘털 셀을 쓰지 않는다 (col 18~21 빈칸)."""
    path = tmp_path / "out.xlsx"
    wb, formats = make_workbook(path)
    failures = [TickerFailure(spec=_spec("XYZ"), reason="네트워크 오류")]
    write_portfolio_sheet(wb, formats, [], failures, ["XYZ"], now=datetime(2026, 5, 26))
    wb.close()

    ws = _open(path)["시트1"]
    for col in (18, 19, 20, 21):
        c = ws.cell(row=6, column=col)
        assert c.value in (None, "")
        assert c.comment is None
