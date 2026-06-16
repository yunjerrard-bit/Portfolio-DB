"""End-to-end smoke + 3-row color verification.

Wave 4 (PLAN 01-05): `stocksig.main.run` (== `stocksig.main_run.run`) 호출 →
실제 XlsxWriter로 .xlsx 생성 → openpyxl로 다시 열어 시트 구조 + 3행 색 검증.

테스트는 mock yfinance fixture(`mock_ohlcv_df`, conftest.py)를 사용하므로
네트워크 호출 없이 결정론적으로 통과한다.
"""

from __future__ import annotations

import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter

from stocksig.compute.color_rules import (
    GREEN_100,
    GREEN_800,
    GREEN_900,
    RED_100,
    RED_800,
    RED_900,
    SigmaBucket,
    decide_sigma_bucket,
)
from stocksig.main import run
from stocksig.output.sheet_per_ticker import build_column_layout

# D-04 hex → openpyxl ARGB (alpha=FF) 매핑
_SIGMA_FONT_HEX = {
    SigmaBucket.DEFAULT: None,
    SigmaBucket.SOFT_GREEN: GREEN_800.lstrip("#").upper(),
    SigmaBucket.SOFT_RED: RED_800.lstrip("#").upper(),
    SigmaBucket.HARD_GREEN: GREEN_900.lstrip("#").upper(),
    SigmaBucket.HARD_RED: RED_900.lstrip("#").upper(),
}
_SIGMA_FILL_HEX = {
    SigmaBucket.DEFAULT: None,
    SigmaBucket.SOFT_GREEN: None,
    SigmaBucket.SOFT_RED: None,
    SigmaBucket.HARD_GREEN: GREEN_100.lstrip("#").upper(),
    SigmaBucket.HARD_RED: RED_100.lstrip("#").upper(),
}


def _normalize_rgb(rgb_value) -> str | None:
    """openpyxl ARGB ('FFRRGGBB') 또는 RGB ('RRGGBB') 또는 None → 'RRGGBB' uppercase 또는 None."""
    if rgb_value is None:
        return None
    if not isinstance(rgb_value, str):
        return None
    v = rgb_value.upper()
    if len(v) == 8:
        return v[2:]
    if len(v) == 6:
        return v
    return None


def _setup_mock_yfinance(mocker, df: pd.DataFrame) -> None:
    """`yf.Ticker(...).history(...)`가 df를 반환하도록 patch.

    Pitfall 5 (06-01): Task 3로 run() 경로에 fetch_company_name 이 추가되면서
    company.py 의 자체 `yf.Ticker` import 가 위 market 패치로 커버되지 않는다.
    네트워크 격리를 위해 `stocksig.io.company.fetch_company_name` 을 결정론적
    stub(티커=이름)으로 patch 한다.
    """
    ticker_class = mocker.patch("stocksig.io.market.yf.Ticker")
    instance = ticker_class.return_value
    instance.history.return_value = df
    # 기업명 fetch 결정론적 stub — 티커를 그대로 이름으로 반환 (네트워크 없음).
    # main_run 은 `from stocksig.io.company import fetch_company_name` 로 이미 바인딩한
    # 참조를 run_all 에 넘기므로, 실제 사용되는 main_run 바인딩을 patch 한다.
    # (원본 정의도 함께 patch — stocksig.io.company 경유 호출 경로까지 격리.)
    mocker.patch(
        "stocksig.io.company.fetch_company_name", side_effect=lambda t: t
    )
    mocker.patch(
        "stocksig.main_run.fetch_company_name", side_effect=lambda t: t
    )


def test_single_ticker_workbook(mocker, mock_ohlcv_df, tmp_path, tmp_tickers_file, tmp_env_file):
    """smoke: AAPL 단일 티커로 워크북 생성 + 시트 구조 검증."""
    _setup_mock_yfinance(mocker, mock_ohlcv_df)
    tickers = tmp_tickers_file("AAPL\n")
    env = tmp_env_file("EDGAR_USER_AGENT_EMAIL=test@example.com\nOPENDART_API_KEY=test-key\n")

    output_path = run(tickers, env, tmp_path / "output")

    assert output_path.exists(), f"워크북이 생성되어야 한다: {output_path}"
    assert output_path.suffix == ".xlsx"

    wb = openpyxl.load_workbook(output_path)
    assert "AAPL" in wb.sheetnames

    # 06-01 HIGH-2: 시트1 기업명 B열 측정 가능 회귀 단언.
    ws1 = wb["시트1"]
    # B5 헤더 셀 == "기업명" (A 티커와 C 시장 사이).
    assert ws1.cell(row=5, column=2).value == "기업명"
    # 데이터행(row 6) B열이 비어있지 않음 — 기업명 또는 티커 폴백이 채워짐.
    assert ws1.cell(row=6, column=2).value not in (None, "")

    ws = wb["AAPL"]

    # SHEET-02: A1 == ticker
    assert ws["A1"].value == "AAPL"

    # 97 컬럼 (gap-fix 01-14 95 + 주봉 EMA 진행형 추세 2개)
    assert ws.max_column == 97, f"시트 폭은 97이어야 한다 (현재: {ws.max_column})"

    # gap-fix 01-07: 신규 한국어 헤더 검증
    layout = build_column_layout()
    diff_high_11_col = layout.index("DIFF_High_11") + 1
    assert ws.cell(row=5, column=diff_high_11_col).value == "고가-EMA11 차이"
    ema_close_11_col = layout.index("EMA_Close_11") + 1
    assert ws.cell(row=5, column=ema_close_11_col).value == "종가 EMA11"

    # SHEET-03: row 3 — Close 위치(col B, index 2) — 스칼라 median 존재
    assert ws.cell(row=3, column=2).value is not None
    assert isinstance(ws.cell(row=3, column=2).value, (int, float))

    # SHEET-04: row 4 — Close 위치 — 스칼라 std 존재
    assert ws.cell(row=4, column=2).value is not None
    assert isinstance(ws.cell(row=4, column=2).value, (int, float))

    # SHEET-05: row 5 한국어 헤더 (gap-fix 01-14: (일) prefix + (200일) suffix)
    assert ws.cell(row=5, column=1).value == "날짜"
    assert ws.cell(row=5, column=2).value == "(일)종가 (200일)"
    assert ws.cell(row=5, column=3).value == "(일)종가 (200일) 일별 중앙값"
    assert ws.cell(row=5, column=4).value == "(일)종가 (200일) 일별 표준편차"

    # SHEET-06: row 6+ 데이터 — 가장 최신 날짜 (descending)
    first_date = ws.cell(row=6, column=1).value
    second_date = ws.cell(row=7, column=1).value
    assert first_date is not None
    assert second_date is not None
    # 첫 행이 두 번째 행보다 더 최신 (내림차순)
    assert first_date > second_date


def test_color_at_three_rows(mocker, mock_ohlcv_df, tmp_path, tmp_tickers_file, tmp_env_file):
    """Success Criteria #2/#3: 3개 행(최신/중간/오래된)의 Close 셀 색이 D-04와 일치.

    expected SigmaBucket은 mock_ohlcv_df로부터 직접 expanding median/std를
    재계산하여 `decide_sigma_bucket`으로 산출 → openpyxl로 읽은 실제 셀 색과 비교.
    """
    _setup_mock_yfinance(mocker, mock_ohlcv_df)
    tickers = tmp_tickers_file("AAPL\n")
    env = tmp_env_file("EDGAR_USER_AGENT_EMAIL=test@example.com\nOPENDART_API_KEY=test-key\n")

    output_path = run(tickers, env, tmp_path / "output")
    wb = openpyxl.load_workbook(output_path)
    ws = wb["AAPL"]

    # gap-fix 01-14: Close median/std 는 rolling(200) — expanding 이 아님
    sorted_asc = mock_ohlcv_df.sort_index(ascending=True)
    close = sorted_asc["Close"]
    expanding_median = close.rolling(window=200, min_periods=200).median()
    expanding_std = close.rolling(window=200, min_periods=200).std()

    n = len(sorted_asc)
    # 3개 인덱스 — gap-fix 01-14: index 0은 rolling(200) NaN 영역이므로 199 이후로 이동
    indices = [199, n // 2, n - 1]

    # 시트는 내림차순 → row 6이 가장 최신 (asc index n-1)
    # row 6+i 가 sorted_asc index n-1-i
    for asc_idx in indices:
        descending_row_offset = (n - 1) - asc_idx  # 0이면 row 6, 1이면 row 7, ...
        excel_row = 6 + descending_row_offset
        value = close.iloc[asc_idx]
        med = expanding_median.iloc[asc_idx]
        std = expanding_std.iloc[asc_idx]
        expected_bucket = decide_sigma_bucket(value, med, std)

        # Close 컬럼은 layout의 index 1 → openpyxl column 2
        cell = ws.cell(row=excel_row, column=2)

        # 값 정합성 (반올림 허용)
        assert cell.value is not None, f"row {excel_row} Close 셀이 비어있다"
        assert abs(float(cell.value) - float(value)) < 1e-6

        # 폰트 색 검증
        expected_font_hex = _SIGMA_FONT_HEX[expected_bucket]
        font_color = cell.font.color
        actual_font_hex = (
            _normalize_rgb(font_color.rgb) if font_color is not None else None
        )
        # 기본(검정/None)은 XlsxWriter가 색을 지정하지 않으므로 None 또는 '000000' 둘 다 허용
        if expected_font_hex is None:
            assert actual_font_hex in (None, "000000"), (
                f"row {excel_row} (bucket={expected_bucket}) 폰트 색이 기본이어야 한다 "
                f"(actual={actual_font_hex})"
            )
        else:
            assert actual_font_hex == expected_font_hex, (
                f"row {excel_row} (bucket={expected_bucket}) 폰트 색 불일치: "
                f"expected={expected_font_hex}, actual={actual_font_hex}"
            )

        # 배경 색 검증 (HARD 케이스만 fill 존재)
        expected_fill_hex = _SIGMA_FILL_HEX[expected_bucket]
        fill = cell.fill
        actual_fill_hex = None
        if fill is not None and fill.fgColor is not None:
            # openpyxl PatternFill: fgColor.type='rgb'일 때만 hex
            if getattr(fill.fgColor, "type", None) == "rgb":
                actual_fill_hex = _normalize_rgb(fill.fgColor.rgb)
        if expected_fill_hex is None:
            assert actual_fill_hex in (None, "000000"), (
                f"row {excel_row} (bucket={expected_bucket}) 배경 색이 없어야 한다 "
                f"(actual={actual_fill_hex})"
            )
        else:
            assert actual_fill_hex == expected_fill_hex, (
                f"row {excel_row} (bucket={expected_bucket}) 배경 색 불일치: "
                f"expected={expected_fill_hex}, actual={actual_fill_hex}"
            )


def test_num_format_baked(mocker, mock_ohlcv_df, tmp_path, tmp_tickers_file, tmp_env_file):
    """gap-fix 01-06: 각 컬럼 타입의 데이터 셀에 num_format이 베이크되어 있어야 한다.

    - 종가 (Close)    → '#,##0.00'
    - 거래량 (Volume) → '#,##0'
    - RSI             → '0.00"%"'
    """
    _setup_mock_yfinance(mocker, mock_ohlcv_df)
    tickers = tmp_tickers_file("AAPL\n")
    env = tmp_env_file("EDGAR_USER_AGENT_EMAIL=test@example.com\nOPENDART_API_KEY=test-key\n")

    output_path = run(tickers, env, tmp_path / "output")
    wb = openpyxl.load_workbook(output_path)
    ws = wb["AAPL"]

    layout = build_column_layout()
    close_col = layout.index("Close") + 1
    volume_col = layout.index("Volume") + 1
    rsi_col = layout.index("RSI") + 1  # gap-fix 01-12: 새 위치 (68번째)
    diff_close_11_col = layout.index("DIFF_Close_11") + 1  # gap-fix 01-07: percent_ratio

    # 데이터 첫 행 = row 6 (내림차순 최신)
    close_cell = ws.cell(row=6, column=close_col)
    volume_cell = ws.cell(row=6, column=volume_col)
    rsi_cell = ws.cell(row=6, column=rsi_col)
    diff_cell = ws.cell(row=6, column=diff_close_11_col)

    assert close_cell.value is not None, "Close 셀이 비어 있다"
    assert volume_cell.value is not None, "Volume 셀이 비어 있다"
    assert rsi_cell.value is not None, "RSI 셀이 비어 있다"
    assert diff_cell.value is not None, "DIFF_Close_11 셀이 비어 있다"

    assert close_cell.number_format == "#,##0.00", (
        f"Close num_format 불일치: expected='#,##0.00', actual={close_cell.number_format!r}"
    )
    assert volume_cell.number_format == "#,##0", (
        f"Volume num_format 불일치: expected='#,##0', actual={volume_cell.number_format!r}"
    )
    assert rsi_cell.number_format == '0.00"%"', (
        f"RSI num_format 불일치: expected='0.00\"%\"', actual={rsi_cell.number_format!r}"
    )
    # gap-fix 01-07: DIFF는 비율 → '0.00%' (Excel가 자동 ×100)
    assert diff_cell.number_format == "0.00%", (
        f"DIFF_Close_11 num_format 불일치: expected='0.00%', actual={diff_cell.number_format!r}"
    )
    # DIFF 값은 비율 스케일 — 절댓값 < 1
    assert abs(float(diff_cell.value)) < 1.0, (
        f"DIFF_Close_11 값이 비율 스케일이 아니다: {diff_cell.value}"
    )


def test_median_std_columns_hidden(mocker, mock_ohlcv_df, tmp_path, tmp_tickers_file, tmp_env_file):
    """gap-fix 01-08: 모든 *_median, *_std 컬럼은 hidden=True, 그 외 가시.

    검증:
      - Close_median, DIFF_Close_11_std 위치 → hidden == True
      - Close, RSI, Stoch_%K 등 primary 컬럼 → hidden == False
      - 데이터·서식은 그대로 유지 (값 존재 확인)
    """
    _setup_mock_yfinance(mocker, mock_ohlcv_df)
    tickers = tmp_tickers_file("AAPL\n")
    env = tmp_env_file("EDGAR_USER_AGENT_EMAIL=test@example.com\nOPENDART_API_KEY=test-key\n")

    output_path = run(tickers, env, tmp_path / "output")
    wb = openpyxl.load_workbook(output_path)
    ws = wb["AAPL"]

    layout = build_column_layout()

    # XlsxWriter는 연속된 set_column 호출을 min/max 범위로 합쳐 저장하므로,
    # openpyxl의 ws.column_dimensions[letter]는 defaultdict 동작으로 빈 항목을 만들 수 있다.
    # 실제 컬럼 속성은 저장된 모든 ColumnDimension 의 (min, max) 범위를 순회해 찾는다.
    def is_hidden(col_excel_idx: int) -> bool:
        for cd in ws.column_dimensions.values():
            if cd.min is None or cd.max is None:
                continue
            if cd.min <= col_excel_idx <= cd.max:
                return bool(cd.hidden)
        return False

    import re as _re
    _EMA_VAL = _re.compile(r"^EMA_Close_\d+$")

    # 숨김 대상 검증 (sample) — gap-fix 01-11/01-12: EMA_Close_N 값 컬럼도 hidden
    hidden_targets = ["Close_median", "Close_std", "DIFF_Close_11_std",
                       "Volume_pct_change_median", "Volume_pct_change_std",
                       "EMA_Close_192_median",
                       "EMA_Close_11", "EMA_Close_192"]
    for col_name in hidden_targets:
        col_idx = layout.index(col_name)
        letter = get_column_letter(col_idx + 1)
        assert is_hidden(col_idx + 1) is True, (
            f"{col_name} (column {letter}) 는 hidden=True 여야 한다"
        )

    # 비숨김 대상 검증 — EMA_Close_N (값)은 이제 hidden이므로 제외, trend는 가시
    # gap-fix 01-12: dailychg 컬럼은 제거되어 검증 대상 아님
    visible_targets = ["Close", "High", "Low", "Volume",
                        "EMA_Close_11_trend", "EMA_Close_192_trend",
                        "DIFF_Close_11", "RSI", "Stoch_%K", "Stoch_%D"]
    for col_name in visible_targets:
        col_idx = layout.index(col_name)
        letter = get_column_letter(col_idx + 1)
        assert is_hidden(col_idx + 1) is False, (
            f"{col_name} (column {letter}) 는 hidden=False 여야 한다"
        )

    # 데이터 유지 확인: 숨겨진 컬럼에도 값이 살아있어야 함
    close_median_idx = layout.index("Close_median")
    close_median_cell = ws.cell(row=6, column=close_median_idx + 1)
    assert close_median_cell.value is not None, "Close_median 셀 데이터는 보존되어야 한다"

    # 전체 layout — *_median / *_std / 정확히 EMA_Close_N 값컬럼이 hidden
    for col_idx, col_name in enumerate(layout):
        letter = get_column_letter(col_idx + 1)
        expected_hidden = (
            col_name.endswith(("_median", "_std")) or bool(_EMA_VAL.match(col_name))
        )
        actual_hidden = is_hidden(col_idx + 1)
        assert actual_hidden == expected_hidden, (
            f"{col_name} (column {letter}) hidden 불일치: "
            f"expected={expected_hidden}, actual={actual_hidden}"
        )


def test_header_freeze_and_colored_bold(
    mocker, mock_ohlcv_df, tmp_path, tmp_tickers_file, tmp_env_file
):
    """gap-fix 01-10: 1) 1~5행 freeze, 2) 색이 칠해진 셀은 bold, DEFAULT 셀은 not bold."""
    _setup_mock_yfinance(mocker, mock_ohlcv_df)
    tickers = tmp_tickers_file("AAPL\n")
    env = tmp_env_file("EDGAR_USER_AGENT_EMAIL=test@example.com\nOPENDART_API_KEY=test-key\n")

    output_path = run(tickers, env, tmp_path / "output")
    wb = openpyxl.load_workbook(output_path)
    ws = wb["AAPL"]

    # (1) freeze_panes: 행 1~5 + A열(날짜) 고정 → openpyxl는 'B6' 문자열로 읽음
    assert ws.freeze_panes == "B6", (
        f"freeze_panes 불일치: expected='B6', actual={ws.freeze_panes!r}"
    )

    # (2) 색 있는 Close 셀과 DEFAULT Close 셀을 찾아 bold 검증
    sorted_asc = mock_ohlcv_df.sort_index(ascending=True)
    close = sorted_asc["Close"]
    # gap-fix 01-14: Close rolling(200)
    expanding_median = close.rolling(window=200, min_periods=200).median()
    expanding_std = close.rolling(window=200, min_periods=200).std()
    n = len(sorted_asc)

    colored_cell_addr: str | None = None
    default_cell_addr: str | None = None
    for asc_idx in range(199, n):
        value = close.iloc[asc_idx]
        med = expanding_median.iloc[asc_idx]
        std = expanding_std.iloc[asc_idx]
        bucket = decide_sigma_bucket(value, med, std)
        descending_row_offset = (n - 1) - asc_idx
        excel_row = 6 + descending_row_offset
        cell = ws.cell(row=excel_row, column=2)  # Close col
        addr = cell.coordinate
        if bucket != SigmaBucket.DEFAULT and colored_cell_addr is None:
            colored_cell_addr = addr
        if bucket == SigmaBucket.DEFAULT and default_cell_addr is None:
            default_cell_addr = addr
        if colored_cell_addr and default_cell_addr:
            break

    assert colored_cell_addr is not None, (
        "mock fixture에서 색 있는 SigmaBucket Close 셀을 찾을 수 없다"
    )
    assert default_cell_addr is not None, (
        "mock fixture에서 DEFAULT SigmaBucket Close 셀을 찾을 수 없다"
    )

    colored_cell = ws[colored_cell_addr]
    default_cell = ws[default_cell_addr]
    assert colored_cell.font.b is True, (
        f"색 있는 셀 {colored_cell_addr}는 bold여야 한다 (actual font.b={colored_cell.font.b})"
    )
    assert not default_cell.font.b, (
        f"DEFAULT 셀 {default_cell_addr}는 not bold여야 한다 (actual font.b={default_cell.font.b})"
    )


def test_diff_columns_ordered_by_period():
    """gap-fix 01-09: DIFF 컬럼이 EMA period 기준으로 그룹화되어야 한다.

    이전(01-07): price 외부, period 내부 → Close_11, Close_22, ..., High_11, ...
    이후(01-09): period 외부, price 내부 → Close_11, High_11, Low_11, Close_22, ...
    """
    layout = build_column_layout()
    diff_cols = [
        c for c in layout
        if c.startswith("DIFF_") and not c.endswith(("_median", "_std"))
    ]
    expected = [
        "DIFF_Close_11", "DIFF_High_11", "DIFF_Low_11",
        "DIFF_Close_22", "DIFF_High_22", "DIFF_Low_22",
        "DIFF_Close_96", "DIFF_High_96", "DIFF_Low_96",
        "DIFF_Close_192", "DIFF_High_192", "DIFF_Low_192",
    ]
    assert diff_cols == expected, f"DIFF 컬럼 순서 불일치: {diff_cols}"

    # median/std siblings 가 각 base 바로 뒤에 interleave 되어 있는지 확인
    for base in expected:
        i = layout.index(base)
        assert layout[i + 1] == f"{base}_median", (
            f"{base} 다음 컬럼이 {base}_median 이어야 함, 실제: {layout[i + 1]}"
        )
        assert layout[i + 2] == f"{base}_std", (
            f"{base} +2 컬럼이 {base}_std 이어야 함, 실제: {layout[i + 2]}"
        )

    # 97 컬럼 (gap-fix 01-14 95 + 주봉 EMA 진행형 추세 2개)
    assert len(layout) == 97, f"총 컬럼 수 97 이어야 함, 실제: {len(layout)}"


def test_ema_value_columns_hidden(mocker, mock_ohlcv_df, tmp_path, tmp_tickers_file, tmp_env_file):
    """gap-fix 01-11: EMA_Close_{11,22,96,192} 값 컬럼은 hidden=True."""
    _setup_mock_yfinance(mocker, mock_ohlcv_df)
    tickers = tmp_tickers_file("AAPL\n")
    env = tmp_env_file("EDGAR_USER_AGENT_EMAIL=test@example.com\nOPENDART_API_KEY=test-key\n")

    output_path = run(tickers, env, tmp_path / "output")
    wb = openpyxl.load_workbook(output_path)
    ws = wb["AAPL"]

    layout = build_column_layout()

    def is_hidden(col_excel_idx: int) -> bool:
        for cd in ws.column_dimensions.values():
            if cd.min is None or cd.max is None:
                continue
            if cd.min <= col_excel_idx <= cd.max:
                return bool(cd.hidden)
        return False

    for n in [11, 22, 96, 192]:
        col_name = f"EMA_Close_{n}"
        col_idx = layout.index(col_name)
        letter = get_column_letter(col_idx + 1)
        assert is_hidden(col_idx + 1) is True, (
            f"{col_name} (column {letter}) 는 hidden=True 여야 한다 (gap-fix 01-11)"
        )
        # 데이터는 그대로 살아 있어야 함
        assert ws.cell(row=6, column=col_idx + 1).value is not None


def test_trend_columns_visible_and_formatted(
    mocker, mock_ohlcv_df, tmp_path, tmp_tickers_file, tmp_env_file
):
    """gap-fix 01-11: EMA_Close_N_trend 컬럼 가시 + num_format == '0.00%' + 한국어 헤더."""
    _setup_mock_yfinance(mocker, mock_ohlcv_df)
    tickers = tmp_tickers_file("AAPL\n")
    env = tmp_env_file("EDGAR_USER_AGENT_EMAIL=test@example.com\nOPENDART_API_KEY=test-key\n")

    output_path = run(tickers, env, tmp_path / "output")
    wb = openpyxl.load_workbook(output_path)
    ws = wb["AAPL"]

    layout = build_column_layout()

    def is_hidden(col_excel_idx: int) -> bool:
        for cd in ws.column_dimensions.values():
            if cd.min is None or cd.max is None:
                continue
            if cd.min <= col_excel_idx <= cd.max:
                return bool(cd.hidden)
        return False

    expected_kr = {
        11: "ema11 추세",
        22: "ema22 추세",
        96: "ema96 추세",
        192: "ema192 추세",
    }
    for n in [11, 22, 96, 192]:
        col_name = f"EMA_Close_{n}_trend"
        col_idx = layout.index(col_name)
        excel_col = col_idx + 1
        assert is_hidden(excel_col) is False, (
            f"{col_name} 는 가시 상태여야 한다 (hidden=False)"
        )
        # 한국어 헤더
        assert ws.cell(row=5, column=excel_col).value == expected_kr[n]
        # num_format: 데이터 행에서 0.00%
        data_cell = ws.cell(row=6, column=excel_col)
        # row 6은 최신, 최신 행은 pct_change non-null이어야 (충분한 mock 데이터 크기)
        assert data_cell.number_format == "0.00%", (
            f"{col_name} num_format 불일치: expected='0.00%', actual={data_cell.number_format!r}"
        )


def test_trend_color_baking(
    mocker, mock_ohlcv_df, tmp_path, tmp_tickers_file, tmp_env_file
):
    """gap-fix 01-11: trend > 0 → GREEN_800 bold, trend < 0 → RED_800 bold."""
    _setup_mock_yfinance(mocker, mock_ohlcv_df)
    tickers = tmp_tickers_file("AAPL\n")
    env = tmp_env_file("EDGAR_USER_AGENT_EMAIL=test@example.com\nOPENDART_API_KEY=test-key\n")

    output_path = run(tickers, env, tmp_path / "output")
    wb = openpyxl.load_workbook(output_path)
    ws = wb["AAPL"]

    layout = build_column_layout()
    # EMA_Close_11_trend 컬럼
    trend_col_idx = layout.index("EMA_Close_11_trend")
    excel_col = trend_col_idx + 1

    # mock_ohlcv_df 에서 EMA_Close_11 + pct_change 직접 재계산
    sorted_asc = mock_ohlcv_df.sort_index(ascending=True)
    ema11 = sorted_asc["Close"].ewm(span=11, adjust=False).mean()
    trend = ema11.pct_change()
    n_rows = len(sorted_asc)

    pos_row = neg_row = None
    pos_value = neg_value = None
    for asc_idx in range(n_rows):
        v = trend.iloc[asc_idx]
        if v is None or (isinstance(v, float) and (v != v)):
            continue
        excel_row = 6 + ((n_rows - 1) - asc_idx)
        if v > 0 and pos_row is None:
            pos_row = excel_row
            pos_value = v
        elif v < 0 and neg_row is None:
            neg_row = excel_row
            neg_value = v
        if pos_row and neg_row:
            break

    assert pos_row is not None, "mock 데이터에서 trend > 0 행이 필요"
    assert neg_row is not None, "mock 데이터에서 trend < 0 행이 필요"

    pos_cell = ws.cell(row=pos_row, column=excel_col)
    neg_cell = ws.cell(row=neg_row, column=excel_col)

    # 값 정합성
    assert abs(float(pos_cell.value) - float(pos_value)) < 1e-9
    assert abs(float(neg_cell.value) - float(neg_value)) < 1e-9

    # 폰트 색 + bold (gap-fix 01-10 컨벤션 동일)
    pos_font_hex = _normalize_rgb(pos_cell.font.color.rgb) if pos_cell.font.color else None
    neg_font_hex = _normalize_rgb(neg_cell.font.color.rgb) if neg_cell.font.color else None
    assert pos_font_hex == GREEN_800.lstrip("#").upper(), (
        f"trend>0 셀 ({pos_cell.coordinate}) 폰트 색 불일치: {pos_font_hex}"
    )
    assert neg_font_hex == RED_800.lstrip("#").upper(), (
        f"trend<0 셀 ({neg_cell.coordinate}) 폰트 색 불일치: {neg_font_hex}"
    )
    assert pos_cell.font.b is True, "trend>0 셀은 bold (01-10 룰)"
    assert neg_cell.font.b is True, "trend<0 셀은 bold (01-10 룰)"


# --- gap-fix 01-13 smoke tests ------------------------------------------------


def _is_hidden_factory(ws):
    def is_hidden(col_excel_idx: int) -> bool:
        for cd in ws.column_dimensions.values():
            if cd.min is None or cd.max is None:
                continue
            if cd.min <= col_excel_idx <= cd.max:
                return bool(cd.hidden)
        return False
    return is_hidden


def test_close_pct_change_visible_and_colored(
    mocker, mock_ohlcv_df, tmp_path, tmp_tickers_file, tmp_env_file
):
    """gap-fix 01-14: Close_pct_change 가시, '0.00%' 포맷, σ bucket 색 (trend 아님)."""
    _setup_mock_yfinance(mocker, mock_ohlcv_df)
    tickers = tmp_tickers_file("AAPL\n")
    env = tmp_env_file("EDGAR_USER_AGENT_EMAIL=test@example.com\nOPENDART_API_KEY=test-key\n")

    output_path = run(tickers, env, tmp_path / "output")
    wb = openpyxl.load_workbook(output_path)
    ws = wb["AAPL"]

    layout = build_column_layout()
    col_idx = layout.index("Close_pct_change")
    excel_col = col_idx + 1
    is_hidden = _is_hidden_factory(ws)
    assert is_hidden(excel_col) is False
    assert ws.cell(row=5, column=excel_col).value == "(일)종가 등락률"

    # 최신 행 (row 6) num_format 확인
    cell = ws.cell(row=6, column=excel_col)
    assert cell.number_format == "0.00%"

    # gap-fix 01-14: σ bucket 색 — 최신 행에서 sigma_bucket 일치 검증
    sorted_asc = mock_ohlcv_df.sort_index(ascending=True)
    cpc = sorted_asc["Close"].pct_change()
    cpc_med = cpc.expanding().median()
    cpc_std = cpc.expanding().std()
    n = len(sorted_asc)
    asc_idx = n - 1
    expected = decide_sigma_bucket(cpc.iloc[asc_idx], cpc_med.iloc[asc_idx], cpc_std.iloc[asc_idx])
    expected_font = _SIGMA_FONT_HEX[expected]
    actual_font = _normalize_rgb(cell.font.color.rgb) if cell.font.color else None
    if expected_font is None:
        assert actual_font in (None, "000000")
    else:
        assert actual_font == expected_font


def test_volume_pct_change_sigma_colored(
    mocker, mock_ohlcv_df, tmp_path, tmp_tickers_file, tmp_env_file
):
    """gap-fix 01-13: Volume_pct_change 가시, '0.00%' 포맷, sigma bucket 색."""
    _setup_mock_yfinance(mocker, mock_ohlcv_df)
    tickers = tmp_tickers_file("AAPL\n")
    env = tmp_env_file("EDGAR_USER_AGENT_EMAIL=test@example.com\nOPENDART_API_KEY=test-key\n")

    output_path = run(tickers, env, tmp_path / "output")
    wb = openpyxl.load_workbook(output_path)
    ws = wb["AAPL"]

    layout = build_column_layout()
    col_idx = layout.index("Volume_pct_change")
    excel_col = col_idx + 1
    is_hidden = _is_hidden_factory(ws)
    assert is_hidden(excel_col) is False
    assert ws.cell(row=5, column=excel_col).value == "(일)거래량 등락률"

    # 데이터 행 num_format 확인 (충분히 뒤쪽 — non-NaN 행)
    data_cell = ws.cell(row=6, column=excel_col)
    assert data_cell.number_format == "0.00%"
    assert data_cell.value is not None

    # mock에서 decide_sigma_bucket 재계산해 적어도 한 행에서 색 일치 검증
    sorted_asc = mock_ohlcv_df.sort_index(ascending=True)
    vpc = sorted_asc["Volume"].pct_change()
    vpc_med = vpc.expanding().median()
    vpc_std = vpc.expanding().std()
    n = len(sorted_asc)

    # 최신 행에서 검증 (asc index n-1 → excel_row 6)
    asc_idx = n - 1
    excel_row = 6
    expected = decide_sigma_bucket(vpc.iloc[asc_idx], vpc_med.iloc[asc_idx], vpc_std.iloc[asc_idx])
    cell = ws.cell(row=excel_row, column=excel_col)
    expected_font = _SIGMA_FONT_HEX[expected]
    actual_font = _normalize_rgb(cell.font.color.rgb) if cell.font.color else None
    if expected_font is None:
        assert actual_font in (None, "000000")
    else:
        assert actual_font == expected_font, (
            f"row {excel_row} (bucket={expected}) 폰트 색 불일치: "
            f"expected={expected_font}, actual={actual_font}"
        )


def test_volume_cell_colored_by_pct_change(
    mocker, mock_ohlcv_df, tmp_path, tmp_tickers_file, tmp_env_file
):
    """gap-fix 01-13: Volume 셀 자체가 Volume_pct_change 부호 기반 trend 색."""
    _setup_mock_yfinance(mocker, mock_ohlcv_df)
    tickers = tmp_tickers_file("AAPL\n")
    env = tmp_env_file("EDGAR_USER_AGENT_EMAIL=test@example.com\nOPENDART_API_KEY=test-key\n")

    output_path = run(tickers, env, tmp_path / "output")
    wb = openpyxl.load_workbook(output_path)
    ws = wb["AAPL"]

    layout = build_column_layout()
    col_idx = layout.index("Volume")
    excel_col = col_idx + 1

    # Volume 셀 num_format 변경 없음
    assert ws.cell(row=6, column=excel_col).number_format == "#,##0"

    sorted_asc = mock_ohlcv_df.sort_index(ascending=True)
    vpc = sorted_asc["Volume"].pct_change()
    n = len(sorted_asc)

    pos_row = neg_row = None
    for asc_idx in range(n):
        v = vpc.iloc[asc_idx]
        if v is None or (isinstance(v, float) and (v != v)):
            continue
        excel_row = 6 + ((n - 1) - asc_idx)
        if v > 0 and pos_row is None:
            pos_row = excel_row
        elif v < 0 and neg_row is None:
            neg_row = excel_row
        if pos_row and neg_row:
            break
    assert pos_row and neg_row

    pos_cell = ws.cell(row=pos_row, column=excel_col)
    neg_cell = ws.cell(row=neg_row, column=excel_col)
    pos_hex = _normalize_rgb(pos_cell.font.color.rgb) if pos_cell.font.color else None
    neg_hex = _normalize_rgb(neg_cell.font.color.rgb) if neg_cell.font.color else None
    assert pos_hex == GREEN_800.lstrip("#").upper()
    assert neg_hex == RED_800.lstrip("#").upper()


def test_volume_median_std_now_pct_change(
    mocker, mock_ohlcv_df, tmp_path, tmp_tickers_file, tmp_env_file
):
    """gap-fix 01-13: 기존 Volume_median/_std 사라지고 Volume_pct_change_median/_std hidden + percent_ratio."""
    _setup_mock_yfinance(mocker, mock_ohlcv_df)
    tickers = tmp_tickers_file("AAPL\n")
    env = tmp_env_file("EDGAR_USER_AGENT_EMAIL=test@example.com\nOPENDART_API_KEY=test-key\n")

    output_path = run(tickers, env, tmp_path / "output")
    wb = openpyxl.load_workbook(output_path)
    ws = wb["AAPL"]

    layout = build_column_layout()
    assert "Volume_median" not in layout
    assert "Volume_std" not in layout
    assert "Volume_pct_change_median" in layout
    assert "Volume_pct_change_std" in layout

    is_hidden = _is_hidden_factory(ws)
    for col_name in ["Volume_pct_change_median", "Volume_pct_change_std"]:
        col_idx = layout.index(col_name)
        excel_col = col_idx + 1
        assert is_hidden(excel_col) is True, f"{col_name} must be hidden"
        # 데이터 행 num_format
        cell = ws.cell(row=6, column=excel_col)
        assert cell.number_format == "0.00%"


# --- gap-fix 01-14 smoke tests -----------------------------------------------


def test_a1_bold_20pt(mocker, mock_ohlcv_df, tmp_path, tmp_tickers_file, tmp_env_file):
    """gap-fix 01-14: A1 셀 = 티커, bold + 20pt."""
    _setup_mock_yfinance(mocker, mock_ohlcv_df)
    tickers = tmp_tickers_file("AAPL\n")
    env = tmp_env_file("EDGAR_USER_AGENT_EMAIL=test@example.com\nOPENDART_API_KEY=test-key\n")

    output_path = run(tickers, env, tmp_path / "output")
    wb = openpyxl.load_workbook(output_path)
    ws = wb["AAPL"]
    a1 = ws["A1"]
    assert a1.value == "AAPL"
    assert a1.font.b is True
    assert a1.font.sz == 20


def test_ohlc_rolling_200(mocker, mock_ohlcv_df, tmp_path, tmp_tickers_file, tmp_env_file):
    """gap-fix 01-14: Close_median rolling(200) — 첫 199행 NaN."""
    _setup_mock_yfinance(mocker, mock_ohlcv_df)
    tickers = tmp_tickers_file("AAPL\n")
    env = tmp_env_file("EDGAR_USER_AGENT_EMAIL=test@example.com\nOPENDART_API_KEY=test-key\n")

    output_path = run(tickers, env, tmp_path / "output")
    wb = openpyxl.load_workbook(output_path)
    ws = wb["AAPL"]

    layout = build_column_layout()
    close_med_idx = layout.index("Close_median") + 1
    n = len(mock_ohlcv_df)
    # 가장 오래된 행 (asc 0) → excel row 6 + (n-1)
    oldest_row = 6 + (n - 1)
    cell_oldest = ws.cell(row=oldest_row, column=close_med_idx)
    assert cell_oldest.value is None

    # asc_idx = 199 → excel row 6 + (n - 1 - 199)
    row_at_199 = 6 + (n - 1 - 199)
    cell_at_199 = ws.cell(row=row_at_199, column=close_med_idx)
    sorted_asc = mock_ohlcv_df.sort_index(ascending=True)
    close = sorted_asc["Close"]
    expected = close.rolling(window=200, min_periods=200).median().iloc[199]
    assert cell_at_199.value is not None
    assert abs(float(cell_at_199.value) - float(expected)) < 1e-6


def test_weekly_columns_present_and_ffilled(
    mocker, mock_ohlcv_df, tmp_path, tmp_tickers_file, tmp_env_file
):
    """gap-fix 01-14: Close_week ffilled — 인접 행에 동일값 존재."""
    _setup_mock_yfinance(mocker, mock_ohlcv_df)
    tickers = tmp_tickers_file("AAPL\n")
    env = tmp_env_file("EDGAR_USER_AGENT_EMAIL=test@example.com\nOPENDART_API_KEY=test-key\n")

    output_path = run(tickers, env, tmp_path / "output")
    wb = openpyxl.load_workbook(output_path)
    ws = wb["AAPL"]

    layout = build_column_layout()
    for col in ["Close_week", "High_week", "Low_week", "Volume_week"]:
        assert col in layout

    cw_idx = layout.index("Close_week") + 1
    vals = [ws.cell(row=6 + i, column=cw_idx).value for i in range(6)]
    vs = [v for v in vals if v is not None]
    assert len(vs) >= 2
    # 적어도 한 쌍의 인접 값이 동일 (forward-fill 결과)
    assert any(vs[i] == vs[i + 1] for i in range(len(vs) - 1))


def test_macd_osc_columns(mocker, mock_ohlcv_df, tmp_path, tmp_tickers_file, tmp_env_file):
    """gap-fix 01-14: MACD_OSC 일/주 finite."""
    _setup_mock_yfinance(mocker, mock_ohlcv_df)
    tickers = tmp_tickers_file("AAPL\n")
    env = tmp_env_file("EDGAR_USER_AGENT_EMAIL=test@example.com\nOPENDART_API_KEY=test-key\n")

    output_path = run(tickers, env, tmp_path / "output")
    wb = openpyxl.load_workbook(output_path)
    ws = wb["AAPL"]

    layout = build_column_layout()
    for col in ["MACD_OSC", "MACD_OSC_week"]:
        idx = layout.index(col) + 1
        cell = ws.cell(row=6, column=idx)
        assert cell.value is not None
        assert isinstance(cell.value, (int, float))


def test_macd_osc_color_by_diff_sign(
    mocker, mock_ohlcv_df, tmp_path, tmp_tickers_file, tmp_env_file
):
    """gap-fix 01-14: MACD_OSC 셀 색 = MACD_OSC.diff() 부호 (trend bucket)."""
    _setup_mock_yfinance(mocker, mock_ohlcv_df)
    tickers = tmp_tickers_file("AAPL\n")
    env = tmp_env_file("EDGAR_USER_AGENT_EMAIL=test@example.com\nOPENDART_API_KEY=test-key\n")

    output_path = run(tickers, env, tmp_path / "output")
    wb = openpyxl.load_workbook(output_path)
    ws = wb["AAPL"]

    layout = build_column_layout()
    idx = layout.index("MACD_OSC") + 1

    from stocksig.compute.indicators import compute_macd_oscillator
    sorted_asc = mock_ohlcv_df.sort_index(ascending=True)
    osc = compute_macd_oscillator(sorted_asc["Close"])
    osc_diff = osc.diff()
    n = len(sorted_asc)

    pos_row = neg_row = None
    for asc_idx in range(50, n):
        v = osc_diff.iloc[asc_idx]
        if v is None or (isinstance(v, float) and (v != v)):
            continue
        excel_row = 6 + ((n - 1) - asc_idx)
        if v > 0 and pos_row is None:
            pos_row = excel_row
        elif v < 0 and neg_row is None:
            neg_row = excel_row
        if pos_row and neg_row:
            break
    assert pos_row and neg_row

    pos_cell = ws.cell(row=pos_row, column=idx)
    neg_cell = ws.cell(row=neg_row, column=idx)
    pos_hex = _normalize_rgb(pos_cell.font.color.rgb) if pos_cell.font.color else None
    neg_hex = _normalize_rgb(neg_cell.font.color.rgb) if neg_cell.font.color else None
    assert pos_hex == GREEN_800.lstrip("#").upper()
    assert neg_hex == RED_800.lstrip("#").upper()


def test_impulse_cells(mocker, mock_ohlcv_df, tmp_path, tmp_tickers_file, tmp_env_file):
    """gap-fix 01-14: Impulse_daily 셀 텍스트 + 색."""
    _setup_mock_yfinance(mocker, mock_ohlcv_df)
    tickers = tmp_tickers_file("AAPL\n")
    env = tmp_env_file("EDGAR_USER_AGENT_EMAIL=test@example.com\nOPENDART_API_KEY=test-key\n")

    output_path = run(tickers, env, tmp_path / "output")
    wb = openpyxl.load_workbook(output_path)
    ws = wb["AAPL"]

    layout = build_column_layout()
    idx = layout.index("Impulse_daily") + 1
    n = len(mock_ohlcv_df)

    expected = {
        "녹색": ("2E7D32", "C8E6C9"),
        "적색": ("C62828", "FFCDD2"),
        "청색": ("1565C0", "BBDEFB"),
    }

    found = False
    for r in range(6, 6 + min(200, n)):
        cell = ws.cell(row=r, column=idx)
        v = cell.value
        if v in expected:
            font_hex = _normalize_rgb(cell.font.color.rgb) if cell.font.color else None
            fill = cell.fill
            fill_hex = None
            if fill is not None and fill.fgColor is not None:
                if getattr(fill.fgColor, "type", None) == "rgb":
                    fill_hex = _normalize_rgb(fill.fgColor.rgb)
            exp_font, exp_fill = expected[v]
            assert font_hex == exp_font
            assert fill_hex == exp_fill
            assert cell.font.b is True
            found = True
            break
    assert found, "Impulse 텍스트 셀을 찾지 못했다"


def test_diff_block_header_bg(mocker, mock_ohlcv_df, tmp_path, tmp_tickers_file, tmp_env_file):
    """gap-fix 01-14: DIFF 4 그룹 헤더 bg 4가지."""
    _setup_mock_yfinance(mocker, mock_ohlcv_df)
    tickers = tmp_tickers_file("AAPL\n")
    env = tmp_env_file("EDGAR_USER_AGENT_EMAIL=test@example.com\nOPENDART_API_KEY=test-key\n")

    output_path = run(tickers, env, tmp_path / "output")
    wb = openpyxl.load_workbook(output_path)
    ws = wb["AAPL"]

    layout = build_column_layout()
    expected = {
        11: "BDD7EE",
        22: "F8CBAD",
        96: "E2EFDA",
        192: "E1BEE7",
    }
    for n, exp_hex in expected.items():
        col_name = f"DIFF_Close_{n}"
        col_idx = layout.index(col_name) + 1
        cell = ws.cell(row=5, column=col_idx)
        fill = cell.fill
        actual = None
        if fill is not None and fill.fgColor is not None:
            if getattr(fill.fgColor, "type", None) == "rgb":
                actual = _normalize_rgb(fill.fgColor.rgb)
        assert actual == exp_hex, (
            f"DIFF_Close_{n} 헤더 bg 불일치: expected={exp_hex}, actual={actual}"
        )


def test_daily_weekly_pairs(mocker, mock_ohlcv_df, tmp_path, tmp_tickers_file, tmp_env_file):
    """gap-fix 01-14: 일/주 쌍 컬럼 + (일)/(주) prefix 헤더."""
    _setup_mock_yfinance(mocker, mock_ohlcv_df)
    tickers = tmp_tickers_file("AAPL\n")
    env = tmp_env_file("EDGAR_USER_AGENT_EMAIL=test@example.com\nOPENDART_API_KEY=test-key\n")

    output_path = run(tickers, env, tmp_path / "output")
    wb = openpyxl.load_workbook(output_path)
    ws = wb["AAPL"]

    layout = build_column_layout()
    pairs = [
        ("Stoch_%K", "Stoch_%K_week"),
        ("Stoch_%D", "Stoch_%D_week"),
        ("RSI", "RSI_week"),
        ("MACD_OSC", "MACD_OSC_week"),
        ("Impulse_daily", "Impulse_weekly"),
    ]
    for d, w in pairs:
        assert d in layout
        assert w in layout
        d_hdr = ws.cell(row=5, column=layout.index(d) + 1).value
        w_hdr = ws.cell(row=5, column=layout.index(w) + 1).value
        assert d_hdr is not None and "(일)" in d_hdr, f"{d} header: {d_hdr}"
        assert w_hdr is not None and "(주)" in w_hdr, f"{w} header: {w_hdr}"
