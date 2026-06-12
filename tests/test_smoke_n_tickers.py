"""Phase 2 Wave 4 (02-04): N-티커 end-to-end smoke (mocked yfinance).

검증 대상 (PLAN 02-04 must_haves):
  - sheetnames[0] == "시트1" (PORT-01)
  - 성공 티커마다 자기 이름의 시트 (입력 순서)
  - 실패 티커는 per-ticker 시트가 없고 시트1에만 `실패: ...` 행으로 표시 (D-03)
  - 같은 날 재실행은 캐시 HIT (MKTD-05) — 2번째 run에서 `cache HIT` 로그
  - 10 티커 mocked 실행이 큰 시간 없이 끝남
  - input_order = [C, A, B] → 시트 순서 = [시트1, C, A, B]
  - 부분 데이터 (rows<1250)는 시트1 실패 행 (`부분 데이터: ...`)

모든 테스트는 `monkeypatch`로 `stocksig.io.market.fetch_ohlcv` stub +
캐시 디렉토리를 `tmp_path`로 격리한다 (네트워크 없음, 결정론적).
"""

from __future__ import annotations

import logging
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
import pytest

import stocksig.io.cache as cache_mod
import stocksig.io.market as market_mod
from stocksig.main_run import run


# --- helpers ---------------------------------------------------------------


def _make_ohlcv(rows: int = 2700, seed: int = 42) -> pd.DataFrame:
    """결정론적 OHLCV — conftest mock_ohlcv_df와 동일 구조."""
    rng = np.random.default_rng(seed=seed)
    dates = pd.date_range(end=pd.Timestamp("2026-05-20"), periods=rows, freq="B")
    drift = rng.normal(loc=0.0, scale=1.0, size=rows).cumsum() * 0.1
    close = 100.0 + drift
    df = pd.DataFrame(
        {
            "Open": close + rng.normal(0.0, 0.5, rows),
            "High": close * 1.02,
            "Low": close * 0.98,
            "Close": close,
            "Volume": rng.integers(1_000_000, 10_000_000, rows),
        },
        index=dates,
    )
    df.index.name = "Date"
    return df


@pytest.fixture
def env_file(tmp_path: Path) -> Path:
    p = tmp_path / ".env"
    p.write_text(
        "EDGAR_USER_AGENT_EMAIL=test@example.com\nOPENDART_API_KEY=test-key\n",
        encoding="utf-8",
    )
    return p


@pytest.fixture
def mock_pipeline_env(monkeypatch, tmp_path):
    """fetch_ohlcv stub + cache dir 격리 + 인증 ping stub — 모든 N-티커 smoke 가 사용."""
    call_counter = {"count": 0, "rows_per": {}}

    # 캐시 격리 — tmp_path/.cache/ohlcv + reset 전역 _cache
    cache_dir = tmp_path / ".cache" / "ohlcv"
    monkeypatch.setattr(cache_mod, "_DEFAULT_DIR", cache_dir)
    monkeypatch.setattr(cache_mod, "_cache", None)

    def _stub_fetch(ticker: str) -> pd.DataFrame:
        call_counter["count"] += 1
        rows = call_counter["rows_per"].get(ticker, 2700)
        if rows == "raise":
            raise RuntimeError(f"{ticker} | mocked failure")
        # 결정론적이지만 ticker별 seed로 살짝 다르게
        seed = abs(hash(ticker)) % 10_000
        return _make_ohlcv(rows=rows, seed=seed)

    monkeypatch.setattr(market_mod, "fetch_ohlcv", _stub_fetch)

    # 인증 ping stub — 네트워크 없이 OK (04-03). 개별 테스트가 monkeypatch 로 재설정 가능.
    import stocksig.main_run as main_mod

    monkeypatch.setattr(main_mod, "ping_edgar", lambda: (True, None))
    monkeypatch.setattr(main_mod, "ping_dart", lambda: (True, None))
    yield call_counter


# --- tests -----------------------------------------------------------------


def test_portfolio_is_first_sheet(mock_pipeline_env, tmp_path, env_file):
    """PORT-01: 시트1 가 sheetnames[0]."""
    tickers = tmp_path / "tickers.txt"
    tickers.write_text("AAPL\nMSFT\nGOOG\n005930.KS\n035720.KQ\n", encoding="utf-8")

    out = run(tickers, env_file, tmp_path / "output")

    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames[0] == "시트1", (
        f"시트1 이 첫 시트여야 한다 (실제: {wb.sheetnames})"
    )


def test_all_tickers_have_sheets(mock_pipeline_env, tmp_path, env_file):
    """성공 티커 5개 + 시트1 = 총 6 시트, 입력 순서 유지."""
    tickers = tmp_path / "tickers.txt"
    tickers.write_text("AAPL\nMSFT\nGOOG\n005930.KS\n035720.KQ\n", encoding="utf-8")

    out = run(tickers, env_file, tmp_path / "output")

    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["시트1", "AAPL", "MSFT", "GOOG", "005930.KS", "035720.KQ"]


def test_failed_ticker_in_sheet1_only(
    mock_pipeline_env, tmp_path, env_file, caplog
):
    """D-03: 실패 티커는 per-sheet 없고 시트1 실패 행에만 등장."""
    mock_pipeline_env["rows_per"]["BAD"] = "raise"
    tickers = tmp_path / "tickers.txt"
    tickers.write_text("AAPL\nBAD\nCSCO\n", encoding="utf-8")

    caplog.set_level(logging.INFO)
    out = run(tickers, env_file, tmp_path / "output")

    wb = openpyxl.load_workbook(out)
    # 실패 티커는 sheet 없음
    assert "BAD" not in wb.sheetnames
    assert wb.sheetnames == ["시트1", "AAPL", "CSCO"]

    # 시트1 에 BAD failure row (마지막 col 에 '실패:' prefix)
    ws = wb["시트1"]
    found = False
    for row in ws.iter_rows(min_row=6, values_only=True):
        if row and row[0] == "BAD":
            # col index 14 (마지막) = '실패: ...'
            assert any(
                isinstance(c, str) and c.startswith("실패:") for c in row if c
            ), f"BAD 행에 '실패:' 메시지 없음: {row}"
            found = True
            break
    assert found, "시트1 에 BAD 실패 행이 없다"


def test_partial_data_marked_failure(
    mock_pipeline_env, tmp_path, env_file
):
    """완화된 임계 (2026-05-26): <20 rows만 실패. 10 rows → 시트1 실패 행 `데이터 부족:`."""
    mock_pipeline_env["rows_per"]["TINY"] = 10  # < 20
    tickers = tmp_path / "tickers.txt"
    tickers.write_text("AAPL\nTINY\n", encoding="utf-8")

    out = run(tickers, env_file, tmp_path / "output")

    wb = openpyxl.load_workbook(out)
    assert "TINY" not in wb.sheetnames
    ws = wb["시트1"]
    found = False
    for row in ws.iter_rows(min_row=6, values_only=True):
        if row and row[0] == "TINY":
            assert any(
                isinstance(c, str) and c.startswith("실패: 데이터 부족")
                for c in row
                if c
            ), f"TINY 행: {row}"
            found = True
            break
    assert found, "TINY (데이터 부족) 실패 행이 시트1 에 없다"


def test_cache_hit_on_second_run(
    mock_pipeline_env, tmp_path, env_file, caplog
):
    """MKTD-05: 같은 날 2번째 run → fetch_ohlcv stub 호출 횟수가 두 배가 아니라 N (캐시 HIT)."""
    tickers = tmp_path / "tickers.txt"
    tickers.write_text("AAPL\nMSFT\nCSCO\n", encoding="utf-8")

    # 1차 — 3 misses → 3 fetch
    run(tickers, env_file, tmp_path / "output1")
    first_run_calls = mock_pipeline_env["count"]
    assert first_run_calls == 3, f"1st run fetch 횟수 {first_run_calls}"

    # 2차 — 모두 cache HIT, fetch 추가 호출 없어야 함
    caplog.clear()
    caplog.set_level(logging.INFO, logger="stocksig.io.cache")
    run(tickers, env_file, tmp_path / "output2")
    assert mock_pipeline_env["count"] == 3, (
        f"2nd run fetch 호출됨 (cache MISS): total={mock_pipeline_env['count']}"
    )
    # caplog 에 cache HIT 가 ticker마다 존재
    hits = [r for r in caplog.records if "cache HIT" in r.getMessage()]
    hit_tickers = {r.getMessage().split(" |")[0] for r in hits}
    assert hit_tickers >= {"AAPL", "MSFT", "CSCO"}, (
        f"cache HIT 로그 누락: {hit_tickers}"
    )


def test_10_tickers_completes(mock_pipeline_env, tmp_path, env_file, caplog):
    """10 티커 (success mix) 가 완료 + 요약 로그 검증."""
    symbols = [
        "AAPL", "MSFT", "GOOG", "AMZN", "META",
        "CSCO", "NVDA", "TSLA", "005930.KS", "035720.KQ",
    ]
    tickers = tmp_path / "tickers.txt"
    tickers.write_text("\n".join(symbols) + "\n", encoding="utf-8")

    caplog.set_level(logging.INFO)
    out = run(tickers, env_file, tmp_path / "output")

    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames[0] == "시트1"
    assert set(wb.sheetnames[1:]) == set(symbols)
    # runner 요약 로그
    msgs = [r.getMessage() for r in caplog.records]
    assert any("총 10 티커 중 성공 10 / 실패 0" in m for m in msgs), (
        f"runner 요약 로그 누락. 최근 메시지: {msgs[-5:]}"
    )


def test_input_order_preserved(mock_pipeline_env, tmp_path, env_file):
    """PORT-02: 시트 순서 = 입력 순서 (as_completed 순서가 아님)."""
    tickers = tmp_path / "tickers.txt"
    # 알파벳 역순 + 한국 종목 사이
    tickers.write_text("CSCO\nAAPL\n005930.KS\nMSFT\n", encoding="utf-8")

    out = run(tickers, env_file, tmp_path / "output")

    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["시트1", "CSCO", "AAPL", "005930.KS", "MSFT"]


def test_summary_block_emitted(mock_pipeline_env, tmp_path, env_file, caplog):
    """EXEC-04 / SC3: run() 종료부에 한국어 실행 요약 블록 + 캐시 통계 + 실패 티커 목록."""
    mock_pipeline_env["rows_per"]["BAD"] = "raise"
    tickers = tmp_path / "tickers.txt"
    tickers.write_text("AAPL\nBAD\nMSFT\n", encoding="utf-8")

    caplog.set_level(logging.INFO)
    run(tickers, env_file, tmp_path / "output")

    msgs = [r.getMessage() for r in caplog.records]
    text = "\n".join(msgs)

    # 요약 블록 헤더
    assert any("실행 요약" in m for m in msgs), f"'실행 요약' 헤더 없음: {msgs[-8:]}"
    # 티커 총/성공/실패 줄
    assert any(
        ("티커: 총" in m and "성공" in m and "실패" in m) for m in msgs
    ), f"티커 요약 줄 없음: {msgs[-8:]}"
    # 캐시 통계 줄 (캐시: + HIT)
    assert any(("캐시:" in m and "HIT" in m) for m in msgs), (
        f"캐시 통계 줄 없음: {msgs[-8:]}"
    )
    # 실패 티커 목록 줄 — BAD 포함
    assert any(("실패 티커:" in m and "BAD" in m) for m in msgs), (
        f"실패 티커 목록 줄 없음: {msgs[-8:]}"
    )

    # 회귀: 기존 per-call cache 로그 + runner 진행 요약이 여전히 출력됨 (대체 아님)
    assert "cache MISS" in text, "기존 per-call cache 로그가 사라짐 (대체됨)"
    assert any("총 3 티커 중 성공 2 / 실패 1" in m for m in msgs), (
        "runner 진행 요약 로그가 사라짐"
    )


def test_summary_block_omits_failure_line_when_no_failures(
    mock_pipeline_env, tmp_path, env_file, caplog
):
    """실패 0건이면 요약 블록 안 '실패 티커:' 줄은 생략된다."""
    tickers = tmp_path / "tickers.txt"
    tickers.write_text("AAPL\nMSFT\n", encoding="utf-8")

    caplog.set_level(logging.INFO)
    run(tickers, env_file, tmp_path / "output")

    msgs = [r.getMessage() for r in caplog.records]
    assert any("실행 요약" in m for m in msgs), "요약 블록은 실패 0건에도 출력됨"
    assert not any("실패 티커:" in m for m in msgs), (
        "실패 0건인데 '실패 티커:' 줄이 출력됨"
    )


def test_scalars_roundtrip_through_attrs(mock_pipeline_env, tmp_path, env_file):
    """T-02-12 mitigation: enriched_df.attrs['scalars'] 가 runner 통과 후 보존됨.

    `_compute_enriched`가 stash 한 scalars dict 가 `runner.run_all`을 거쳐
    PASS 2 단계에서도 retrieve 가능해야 한다 (pandas `attrs` 보존성 검증).
    """
    tickers = tmp_path / "tickers.txt"
    tickers.write_text("AAPL\n", encoding="utf-8")

    # 직접 pipeline + run_all 경로로 attrs 검증
    from stocksig.io.input import read_tickers_extended
    from stocksig.io.market_kind import classify_market
    from stocksig.main_run import _make_pipeline
    from stocksig.runner import run_all

    # 워크북 path 만들지 않고 PASS 1만 수행
    specs = read_tickers_extended(tickers)
    results, failures = run_all(specs, classify_market, _make_pipeline())
    assert len(results) == 1
    assert len(failures) == 0

    scalars = results[0].enriched_df.attrs.get("scalars")
    assert scalars is not None, "scalars attrs 가 runner 통과 후 사라짐"
    assert "Close" in scalars
    assert "median" in scalars["Close"]
    assert "std" in scalars["Close"]
    assert isinstance(scalars["Close"]["median"], float)


# --- 04-03 Task 3: 조건부 인증 ping 배선 + skip 클로저 + 요약 인증 줄 -------


def test_ping_edgar_called_when_us_ticker(mock_pipeline_env, tmp_path, env_file, monkeypatch):
    """US 티커 포함 시 ping_edgar 호출됨."""
    import stocksig.main_run as main_mod

    calls = {"edgar": 0, "dart": 0}
    monkeypatch.setattr(main_mod, "ping_edgar", lambda: (calls.__setitem__("edgar", calls["edgar"] + 1) or (True, None)))
    monkeypatch.setattr(main_mod, "ping_dart", lambda: (calls.__setitem__("dart", calls["dart"] + 1) or (True, None)))

    tickers = tmp_path / "tickers.txt"
    tickers.write_text("AAPL\n005930.KS\n", encoding="utf-8")
    run(tickers, env_file, tmp_path / "output")

    assert calls["edgar"] == 1
    assert calls["dart"] == 1


def test_ping_conditional_us_only(mock_pipeline_env, tmp_path, env_file, monkeypatch):
    """D-04 조건부: US 티커만 있으면 ping_dart 미호출."""
    import stocksig.main_run as main_mod

    calls = {"edgar": 0, "dart": 0}
    monkeypatch.setattr(main_mod, "ping_edgar", lambda: (calls.__setitem__("edgar", calls["edgar"] + 1) or (True, None)))
    monkeypatch.setattr(main_mod, "ping_dart", lambda: (calls.__setitem__("dart", calls["dart"] + 1) or (True, None)))

    tickers = tmp_path / "tickers.txt"
    tickers.write_text("AAPL\nMSFT\n", encoding="utf-8")
    run(tickers, env_file, tmp_path / "output")

    assert calls["edgar"] == 1
    assert calls["dart"] == 0


def test_ping_conditional_kr_only(mock_pipeline_env, tmp_path, env_file, monkeypatch):
    """D-04 조건부: KR 티커만 있으면 ping_edgar 미호출."""
    import stocksig.main_run as main_mod

    calls = {"edgar": 0, "dart": 0}
    monkeypatch.setattr(main_mod, "ping_edgar", lambda: (calls.__setitem__("edgar", calls["edgar"] + 1) or (True, None)))
    monkeypatch.setattr(main_mod, "ping_dart", lambda: (calls.__setitem__("dart", calls["dart"] + 1) or (True, None)))

    tickers = tmp_path / "tickers.txt"
    tickers.write_text("005930.KS\n035720.KQ\n", encoding="utf-8")
    run(tickers, env_file, tmp_path / "output")

    assert calls["edgar"] == 0
    assert calls["dart"] == 1


def test_ping_failure_continues_and_builds_workbook(
    mock_pipeline_env, tmp_path, env_file, monkeypatch
):
    """D-02: ping_edgar 가 (False, note) 여도 run() 이 예외 없이 워크북 생성."""
    import stocksig.main_run as main_mod

    monkeypatch.setattr(main_mod, "ping_edgar", lambda: (False, "EDGAR 403 (UA 확인)"))

    tickers = tmp_path / "tickers.txt"
    tickers.write_text("AAPL\nMSFT\n", encoding="utf-8")
    out = run(tickers, env_file, tmp_path / "output")

    assert out.exists()
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames[0] == "시트1"


def test_ping_failure_propagates_skip_edgar(
    mock_pipeline_env, tmp_path, env_file, monkeypatch
):
    """ping 실패 시 US 종목 펀더멘털 fetch 에 skip_edgar=True 전달(클로저 경유)."""
    import stocksig.main_run as main_mod

    monkeypatch.setattr(main_mod, "ping_edgar", lambda: (False, "EDGAR 인증 실패"))

    captured = {"calls": []}

    def _spy_fund(ticker, market, last_close, **kwargs):
        captured["calls"].append((ticker, market, kwargs.get("skip_edgar"), kwargs.get("skip_dart")))
        return main_mod.fetch_fundamentals(ticker, market, last_close, **kwargs)

    monkeypatch.setattr(main_mod, "fetch_fundamentals", _spy_fund)

    tickers = tmp_path / "tickers.txt"
    tickers.write_text("AAPL\n", encoding="utf-8")
    run(tickers, env_file, tmp_path / "output")

    us_calls = [c for c in captured["calls"] if c[1] == "US"]
    assert us_calls, f"US 펀더멘털 호출 없음: {captured['calls']}"
    assert all(c[2] is True for c in us_calls), f"skip_edgar 전파 실패: {us_calls}"


def test_summary_has_auth_line(mock_pipeline_env, tmp_path, env_file, caplog):
    """종료부 요약 블록에 '인증:' + EDGAR + DART 줄이 출력됨."""
    tickers = tmp_path / "tickers.txt"
    tickers.write_text("AAPL\n005930.KS\n", encoding="utf-8")

    caplog.set_level(logging.INFO)
    run(tickers, env_file, tmp_path / "output")

    msgs = [r.getMessage() for r in caplog.records]
    assert any(
        ("인증:" in m and "EDGAR" in m and "DART" in m) for m in msgs
    ), f"인증 요약 줄 없음: {msgs[-8:]}"


def test_summary_auth_line_no_secret_leak(
    mock_pipeline_env, tmp_path, env_file, caplog, monkeypatch
):
    """보안: 인증 줄에 OPENDART_API_KEY 값/EDGAR UA 이메일이 포함되지 않음."""
    import stocksig.main_run as main_mod

    # ping 이 실패 사유를 돌려줘도(_auth_label 이 note 사용) 키/UA 가 새지 않아야 함.
    monkeypatch.setattr(main_mod, "ping_edgar", lambda: (False, "EDGAR 인증 실패"))
    monkeypatch.setattr(main_mod, "ping_dart", lambda: (False, "DART 인증 실패"))

    secret_email = "yunjerrard@gmail.com"
    secret_key = "supersecretdartkey123456"
    env_p = tmp_path / ".env"
    env_p.write_text(
        f"EDGAR_USER_AGENT_EMAIL={secret_email}\nOPENDART_API_KEY={secret_key}\n",
        encoding="utf-8",
    )

    tickers = tmp_path / "tickers.txt"
    tickers.write_text("AAPL\n005930.KS\n", encoding="utf-8")

    caplog.set_level(logging.INFO)
    run(tickers, env_p, tmp_path / "output")

    auth_lines = [r.getMessage() for r in caplog.records if "인증:" in r.getMessage()]
    assert auth_lines, "인증 줄이 없음"
    for line in auth_lines:
        assert secret_email not in line, f"UA 이메일 누설: {line}"
        assert secret_key not in line, f"DART 키 누설: {line}"
